# Builds vsyst-cmo-workbook.xlsx — the CMO Toolkit workbook.
# Styling mirrors vsyst-ceo-workbook.xlsx / vsyst-coo-workbook.xlsx exactly.
# Tabs are generated from the CMO-Docs toolkit templates M03, M07, M09, M11, M12, M13
# and lessons 05 §6 / 11 §7.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

INK, MUTED, HDR_BG = "0F172A", "64748B", "334155"
EX_FILL = PatternFill("solid", fgColor="FFF9C4")   # example rows (grey italic on pale yellow)
FM_FILL = PatternFill("solid", fgColor="E8F0FE")   # computed / formula cells
SEC_FILL = PatternFill("solid", fgColor="E2E8F0")  # section band
IN_FILL = PatternFill("solid", fgColor="DCFCE7")   # input cells you type into
HDR_FILL = PatternFill("solid", fgColor=HDR_BG)
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
AMB_FILL = PatternFill("solid", fgColor="FEF3C7")
GRN_FILL = PatternFill("solid", fgColor="DCFCE7")
RS = '"₹"#,##0;-"₹"#,##0'
RS2 = '"₹"#,##0.00;-"₹"#,##0.00'
PCT = '0.0%'
DT = "dd mmm yyyy"
THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BUILT = "2026-08-27"

wb = openpyxl.Workbook()
wb.remove(wb.active)


def sheet(name, title, sub, widths, tab="475569"):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    ws["A1"] = title
    ws["A1"].font = Font(sz=13, b=True, color=INK); ws["A1"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 23
    ws["A2"] = sub
    ws["A2"].font = Font(sz=9, i=True, color=MUTED); ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 26
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def band(ws, row, text, span):
    ws.cell(row=row, column=1, value=text).font = Font(sz=10, b=True, color=INK)
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SEC_FILL


def header(ws, row, cols):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(sz=10, b=True, color="FFFFFF"); c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30


def put(ws, row, vals, start=1, example=False, fmt=None, formula_cols=(), input_cols=()):
    for i, v in enumerate(vals, start=start):
        if v is None: continue
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        if example:
            c.font = Font(i=True, color="808080")
            c.fill = EX_FILL
        if i in formula_cols: c.fill = FM_FILL
        if i in input_cols and not example: c.fill = IN_FILL
        if fmt and i in fmt: c.number_format = fmt[i]
        c.alignment = Alignment(wrap_text=True, vertical="top")


def dv_list(ws, cells, options, title="Pick from list"):
    d = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                       allow_blank=True, showErrorMessage=True)
    d.errorTitle = title
    d.error = "Please pick a value from the dropdown."
    ws.add_data_validation(d)
    d.add(cells)
    return d


def dv_int(ws, cells, lo, hi, title="1–5 only"):
    d = DataValidation(type="whole", operator="between", formula1=str(lo), formula2=str(hi),
                       allow_blank=True, showErrorMessage=True)
    d.errorTitle = title
    d.error = f"Score this factor {lo}–{hi}. See the rubric on this sheet."
    ws.add_data_validation(d)
    d.add(cells)
    return d


def note(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(sz=9, i=True, color=MUTED)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 26


# ─────────────────────────────────────────────── README
ws = sheet("README", "README — what each tab does, who owns it, and how to fill it",
  f"VSYST CMO Workbook · version 1.0 · built {BUILT} · companion to CMO-Docs. "
  "Grey italic on pale yellow = illustrative example data, overwrite it. Pale blue = a live formula, do not type over it. "
  "Pale green = a cell you are meant to type into. Import to Google Sheets: File → Import → Upload → Replace spreadsheet, "
  "then re-check the formulas on the CAC & Payback and Pipeline tabs.",
  [5, 24, 54, 54, 30], tab=MUTED)
header(ws, 4, ["#", "Tab", "What it does", "Key formulas / how to fill", "Owner · cadence"])
ws.freeze_panes = "A5"
readme = [
 ("README", "This page: tab guide, colour legend, the VERIFY LIVE rule, Google Sheets import note.",
  "Read the legend below before touching anything. Every tab names the toolkit template it comes from.",
  "CEO · on every version bump"),
 ("Target List", "The Raipur 100 — a hundred NAMED pumps with a route to an introduction. Template M03. It is the denominator for HC3 (~60% of ICP pumps within 60 km) and the source of every row in Pipeline.",
  "Score the four rubric factors 1–5 in columns J–M; Priority score (N) = J×3 + K×2 + L×2 + M×3, max 50. Priority band (O) computes A ≥38 · B 26–37 · C ≤25. No row without an owner's name; no row without a dated next action.",
  "Domain director · 30 rows in Week 1, 100 by Week 4, groomed every Monday"),
 ("Pipeline", "Every named pump's stage, next action and objection. Template M07. Nine stages, each with an exit test; the conversion block computes your own funnel rates off your own rows.",
  "Stage is a dropdown — do not add a stage. Days since last touch (G) = TODAY() − Last touch; over 14 days turns red and gets a decision on Monday. The conversion block counts rows that REACHED each stage, so it reads a snapshot correctly.",
  "Domain director · same evening as every visit; reviewed Monday 10 min"),
 ("Marketing Scorecard", "The eight weekly marketing numbers (HC8), thirteen weeks trailing. Template M13 §A. Deliberately the same shape as the COO's T05 KPI Scorecard.",
  "Fill weeks left to right with no gaps. Last (V) = last filled week; Trend (W) compares it with the week before; Status (X) computes Green/Amber/Red from Direction, Goal and Red line. Eight rows, no more.",
  "One named owner per row · Monday by 09:30"),
 ("CAC & Payback", "What each channel actually cost, per paying dealer, and how many months of gross profit it takes to earn it back. Template M13 §B–D.",
  "Total channel cost = cash + founder hours × imputed ₹/hour + travel. CAC = total ÷ new paying dealers. Payback = CAC ÷ (ARPA × gross margin), plus any free months given. Blended CAC is a sanity check, never a decision.",
  "CEO · rows quarterly, review the first Monday monthly"),
 ("Campaign Budget", "One campaign on one page: brief, budget planned vs actual, the weekly log, the post-mortem and the multi-district block. Template M12. Filled here with the Raipur 100 example.",
  "Variance ₹ = actual − planned (positive is overspend, and turns red). The weekly log totals itself. Post-mortem CAC = spend ÷ activated. The T+30 district gate computes Go ≥5 · Watch 3–4 · Stop <3.",
  "Third director · one brief per campaign; log Monday 09:30; post-mortem within 5 days"),
 ("Content Calendar", "One piece a week — what, who shoots it, who edits it, where it went, and whether it earned a reply. Template M11 §D.",
  "Reply rate = replies within 48h ÷ recipients; it is scorecard row 7. The 4-week rate at the bottom is Σ replies ÷ Σ recipients, not an average of percentages. Zero replies twice = the row is deleted, not debated.",
  "CEO (site) · domain director (WhatsApp, shoots) · third director (edits) · row filled Friday, broadcast Tuesday"),
 ("Referral Tracker", "Every Dealer Dost ask, including the ones that failed, and what is owed. Template M09 §C.",
  "Reward due (₹) computes from Stage: ₹0 until Live, ₹2,000 at Live, ₹5,000 once they are Paying in month 3 — all VERIFY LIVE against the signed tier sheet. Totals block flags the ₹20,000/year §194R TDS line.",
  "Domain director · logged the same day; reviewed Friday, 30 min"),
 ("Channel Experiments", "One channel, two weeks, one number, a verdict. Lesson 05 §6, reviewed monthly per lesson 11 §7.",
  "Total cost = cash + hours × the imputed rate on CAC & Payback + travel. CAC = total cost ÷ paying dealers. Kill line is written BEFORE the start date; the verdict cell stays empty until the end date.",
  "Third director · one row per experiment; verdict on the end date"),
]
r = 5
for i, (tab, what, how, own) in enumerate(readme, start=1):
    put(ws, r, [i, tab, what, how, own])
    ws.row_dimensions[r].height = 58
    r += 1
r += 1
ws.cell(row=r, column=1, value="Colour legend").font = Font(sz=10, b=True, color=INK); r += 1
for fill, label in [
        (EX_FILL, "Illustrative example — grey italic on pale yellow. Every one of these rows is fictional: invented pumps, invented owners, invented numbers. Overwrite them; they are there to show the shape."),
        (FM_FILL, "Computed cell — a live formula. Do not type over it."),
        (IN_FILL, "Input cell — this is where you type."),
        (SEC_FILL, "Section band — a heading inside the sheet.")]:
    ws.cell(row=r, column=1).fill = fill; ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=2, value=label).font = Font(sz=9, color=MUTED)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 26
    r += 1
r += 1
ws.cell(row=r, column=1, value="VERIFY LIVE").font = Font(sz=10, b=True, color="B91C1C"); r += 1
note(ws, r,
     "Every rupee figure in this workbook is illustrative or was true on " + BUILT + ". "
     "The plan tiers (₹599 / ₹1,799 / ₹4,999 + Enterprise, ex-GST, per GSTIN, billing web-only, 14-day no-card trial) are UNDER OWNER SIGN-OFF — never quote them as final, and never quote the older ₹999–2,499 hybrid numbers. "
     "The ₹400/hour imputed founder rate, the ₹2,000 / ₹3,000 Dealer Dost bounties, the ₹30–50k monthly envelope, the WhatsApp per-message rates and every CAC, goal and red line on these tabs are VERIFY LIVE: "
     "confirm with the CEO (pricing), the CA (TDS §194R, GST on WhatsApp fees) or the BSP (message rates) before a decision turns on them.", 5)
ws.row_dimensions[r].height = 74
r += 2
note(ws, r,
     "Where a tab differs from its template: Target List splits M03's 'Location · km' and 'Next action + date' into separate columns and adds the four rubric factor columns the score is computed from; "
     "Pipeline splits M07's 'Next action + date' and 'Referral asked?' the same way and adds a computed 'Days since last touch'; "
     "Content Calendar splits M11's 'Replies / Views @48h' into recipients, replies, computed reply rate and views. "
     "Nothing else was added, removed or reordered.", 5)
ws.row_dimensions[r].height = 60

# ─────────────────────────────────────────────── Target List
ws = sheet("Target List", "Target List — the Raipur 100, named, scored and routed to an introduction",
  "Fills M03 and exercise 8.3 in lesson 03. A row is a pump, not a company. No row without the owner's name; no row without a dated next action. "
  "Estimated fields carry 'est.' until the first visit. Disqualified rows stay, status DQ, with a re-look date. Re-score after every visit — pump #3 changes pump #11.",
  [30, 8, 20, 9, 18, 11, 11, 9, 30, 8, 8, 8, 8, 10, 8, 15, 34, 12, 14])
band(ws, 4, "A. The list — 100 named pumps within 60 km of Raipur", 19)
header(ws, 5, ["Pump name", "OMC", "Location", "km from Raipur", "Owner name", "Second-gen?",
               "Est. B2B credit customers", "Credit volume band", "Intro path (who introduces us, by name)",
               "Credit volume 1-5 (×3)", "Second-gen 1-5 (×2)", "Distance 1-5 (×2)", "Intro strength 1-5 (×3)",
               "Priority score /50", "Priority band", "Status", "Next action", "Next date", "Owner (ours)"])
ws.freeze_panes = "B6"
TL_FIRST = 6
tl_ex = [
 ("Maa Danteshwari Fuels (fictional)", "IOCL", "Tatibandh", 8, "R. Verma", "Y", "24 est.", "A",
  "Domain director — school friend", 5, 5, 5, 5, "Intro requested", "Call Verma-ji, ask for a Tuesday slot", date(2026, 9, 2), "Domain dir."),
 ("Highway Auto Centre (fictional)", "BPCL", "NH-53, Kharora", 34, "S. Sahu", "Y", "30 est.", "A",
  "Transporter who fuels there", 5, 5, 3, 3, "New", "Visit on the Kharora cluster run", date(2026, 9, 5), "Domain dir."),
 ("Shree Balaji Filling (fictional)", "HPCL", "Bhanpuri", 11, "M. Agrawal", "N", "12 est.", "B",
  "Our CA handles his books", 3, 1, 5, 3, "Researching", "CA to introduce over WhatsApp", date(2026, 9, 3), "CEO"),
 ("Ring Road Fuels (fictional)", "IOCL", "Ring Road 2", 6, "K. Dewangan", "unknown", "4 est.", "C",
  "none yet", 1, 2, 5, 1, "New", "Walk-in during the Tatibandh cluster run", date(2026, 9, 9), "Third dir."),
 ("City Point Petroleum (fictional)", "BPCL", "Pandri", 4, "(manager-run, owner in Dubai)", "N", "0", "C",
  "—", 1, 1, 5, 1, "DQ", "Re-look — owner may return", date(2027, 3, 1), "—"),
]
r = TL_FIRST
for row in tl_ex:
    put(ws, r, list(row[:13]) + [
        f"=IF(COUNT(J{r}:M{r})<4,\"\",J{r}*3+K{r}*2+L{r}*2+M{r}*3)",
        f"=IF(N{r}=\"\",\"\",IF(N{r}>=38,\"A\",IF(N{r}>=26,\"B\",\"C\")))"] + list(row[13:]),
        example=True, fmt={4: "0", 14: "0", 18: DT}, formula_cols=(14, 15))
    r += 1
for rr in range(r, r + 35):
    put(ws, rr, [""] * 9 + ["", "", "", "",
        f"=IF(COUNT(J{rr}:M{rr})<4,\"\",J{rr}*3+K{rr}*2+L{rr}*2+M{rr}*3)",
        f"=IF(N{rr}=\"\",\"\",IF(N{rr}>=38,\"A\",IF(N{rr}>=26,\"B\",\"C\")))", "", "", "", ""],
        fmt={4: "0", 14: "0", 18: DT}, formula_cols=(14, 15),
        input_cols=(1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17, 18, 19))
TL_LAST = r + 34
dv_list(ws, f"B{TL_FIRST}:B{TL_LAST}", ["IOCL", "BPCL", "HPCL", "Reliance", "Nayara", "Shell", "other"])
dv_list(ws, f"F{TL_FIRST}:F{TL_LAST}", ["Y", "N", "unknown"])
dv_list(ws, f"H{TL_FIRST}:H{TL_LAST}", ["A", "B", "C"])
dv_list(ws, f"P{TL_FIRST}:P{TL_LAST}",
        ["New", "Researching", "Intro requested", "Visited", "Demo", "Trial", "Won", "Lost", "DQ"])
dv_list(ws, f"S{TL_FIRST}:S{TL_LAST}", ["CEO", "Domain dir.", "Third dir.", "—"])
dv_int(ws, f"J{TL_FIRST}:M{TL_LAST}", 1, 5)
ws.conditional_formatting.add(f"N{TL_FIRST}:N{TL_LAST}", ColorScaleRule(
    start_type="num", start_value=10, start_color="FEE2E2",
    mid_type="num", mid_value=30, mid_color="FEF3C7",
    end_type="num", end_value=50, end_color="DCFCE7"))
ws.conditional_formatting.add(f"R{TL_FIRST}:R{TL_LAST}", CellIsRule(
    operator="lessThan", formula=["TODAY()"], fill=RED_FILL))
r = TL_LAST + 2
band(ws, r, "B. The scoring rubric — four factors, weighted, maximum 50", 19); r += 1
header(ws, r, ["Factor", "Weight", "5", "3", "1", "", "", "", "", "", "", "", "", "", "", "", "Rule", "", ""]); r += 1
rub = [("B2B credit volume", "×3", "20+ credit customers, band A", "8–15 customers, band B", "Cash/UPI only",
        "The one factor worth arguing about — it is the whole reason he would pay"),
       ("Second-generation presence", "×2", "Son/daughter running operations", "Involved but not deciding", "Owner 65+, nobody younger",
        "The son influences; the father signs (HC5)"),
       ("Distance from Raipur", "×2", "≤20 km", "21–40 km", "60 km+",
        "Straight-line km; drives the cluster you visit in one morning"),
       ("Intro path strength", "×3", "The domain director knows him", "A CA or association contact", "None yet",
        "The one column nobody can scrape")]
for f, w, five, three, one, rule in rub:
    put(ws, r, [f, w, five, three, one] + [None] * 11 + [rule], example=True)
    ws.row_dimensions[r].height = 26
    r += 1
r += 1
put(ws, r, ["Bands", "A ≥ 38", "B 26–37", "C ≤ 25", "Work A first, in geographic clusters, so one morning covers four pumps."],
    example=True); r += 2
band(ws, r, "C. Monday grooming — 20 minutes before the ops meeting", 19); r += 1
header(ws, r, ["Check", "Count", "Rule", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
checks = [
 ("Rows on the list", f"=SUMPRODUCT(1*(A{TL_FIRST}:A{TL_LAST}<>\"\"))", "30 by Week 1, 100 by Week 4. HC3 needs ~60% of ICP pumps within 60 km, and you cannot claim a percentage without a count."),
 ("Band A rows", f"=COUNTIF(O{TL_FIRST}:O{TL_LAST},\"A\")", "These are this week's visits. Cluster them geographically."),
 ("Band B rows", f"=COUNTIF(O{TL_FIRST}:O{TL_LAST},\"B\")", "Worked after the A rows, or when a cluster run passes the door."),
 ("Band C rows", f"=COUNTIF(O{TL_FIRST}:O{TL_LAST},\"C\")", "Not visits. To visit one, change a factor score and say which."),
 ("Rows with no owner name", f"=SUMPRODUCT((A{TL_FIRST}:A{TL_LAST}<>\"\")*(E{TL_FIRST}:E{TL_LAST}=\"\"))", "An address is not a lead. Must be zero before scoring."),
 ("Rows with no dated next action", f"=SUMPRODUCT((A{TL_FIRST}:A{TL_LAST}<>\"\")*(R{TL_FIRST}:R{TL_LAST}=\"\"))", "No date, no action, row is dead."),
 ("Next actions past due", f"=SUMPRODUCT((R{TL_FIRST}:R{TL_LAST}<>\"\")*(R{TL_FIRST}:R{TL_LAST}<TODAY()))", "Every one gets a new date or the row becomes DQ on Monday."),
 ("Disqualified (DQ)", f"=COUNTIF(P{TL_FIRST}:P{TL_LAST},\"DQ\")", "DQ rows STAY, with a re-look date — or you research them again in three months."),
]
for n, f, rule in checks:
    put(ws, r, [n, f, rule], formula_cols=(2,), fmt={2: "#,##0"})
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=19)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
note(ws, r,
     "Scores on the five example rows are computed strictly from the rubric above, so they differ by a few points from the illustrative totals printed in M03's worked-example table "
     "(that table's numbers were written by hand and are not reproducible from its own rubric — 47/42/34/21/12 there, 50/40/30/20/18 here). The ranking, which is the point of the sheet, is identical. "
     "Sources for building the list: IOCL / BPCL / HPCL outlet locators, PPAC's state-wise retail-outlet file, association rolls, hardware service engineers, and the domain director's phone — see M03.", 19)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── Pipeline
ws = sheet("Pipeline", "Pipeline — every named pump, its stage, its next action and the objection actually blocking it",
  "Fills M07 and lesson 06. Nine stages, each with an exit test (section B) — do not add a stage. A stage moves only on its exit test; a cold row becomes Lost with a reason, never quietly Demoed again. "
  "Objections are recorded VERBATIM, in his language — a paraphrase feeds nothing back into M05. Plan / tier and MRR are VERIFY LIVE, owner sign-off pending.",
  [28, 16, 14, 13, 16, 12, 10, 32, 12, 32, 12, 12, 11, 10, 11, 12, 12])
band(ws, 4, "A. The pipeline", 17)
header(ws, 5, ["Pump", "Owner (dealer)", "Segment", "Source channel", "Stage", "Last touch",
               "Days since last touch", "Next action (a verb and a name)", "Next date", "Objection heard (verbatim)",
               "Trial start", "Activation date", "Plan / tier", "MRR ₹", "Referral asked?", "Referral asked on", "Owner (ours)"])
ws.freeze_panes = "B6"
PL_FIRST = 6
pl_ex = [
 ("Maa Danteshwari Fuels", "R. Verma", "pump dealer", "referral", "Paying", date(2026, 8, 25),
  "Activation check at his month-end — Domain dir.", date(2026, 8, 29), "—",
  date(2026, 8, 5), None, "₹1,799", 1799, "Y", date(2026, 8, 25), "Domain dir."),
 ("Highway Auto Centre", "S. Sahu", "pump dealer", "OMC TM", "Trial started", date(2026, 8, 26),
  "Day-14 meeting with the munim present — Domain dir.", date(2026, 9, 9), "\"munim ko sikhana padega\"",
  date(2026, 8, 26), None, "—", None, "N", None, "Domain dir."),
 ("Shree Balaji Filling", "M. Agrawal", "pump dealer", "CA", "Demoed", date(2026, 8, 21),
  "Re-demo with the son present — CEO", date(2026, 9, 2), "\"beta dekhega\"",
  None, None, "—", None, "N", None, "CEO"),
 ("Ring Road Fuels", "K. Dewangan", "pump dealer", "field visit", "Visited", date(2026, 8, 8),
  "", None, "\"abhi Tally theek hai\"",
  None, None, "—", None, "N", None, "Third dir."),
 ("City Point Petroleum", "(manager-run)", "pump dealer", "field visit", "Lost", date(2026, 8, 4),
  "—", None, "\"malik Dubai mein hain\"",
  None, None, "—", None, "N", None, "—"),
]
r = PL_FIRST
for row in pl_ex:
    put(ws, r, list(row[:6]) + [f"=IF(F{r}=\"\",\"\",TODAY()-F{r})"] + list(row[6:]),
        example=True, fmt={6: DT, 7: "0", 9: DT, 11: DT, 12: DT, 14: RS, 16: DT}, formula_cols=(7,))
    ws.row_dimensions[r].height = 30
    r += 1
for rr in range(r, r + 40):
    put(ws, rr, [""] * 6 + [f"=IF(F{rr}=\"\",\"\",TODAY()-F{rr})"] + [""] * 10,
        fmt={6: DT, 7: "0", 9: DT, 11: DT, 12: DT, 14: RS, 16: DT}, formula_cols=(7,),
        input_cols=(1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17))
PL_LAST = r + 39
STAGES = ["Target", "Contacted", "Visited", "Demoed", "Trial started", "Paying",
          "Activated (both sides)", "Referring", "Lost"]
dv_list(ws, f"E{PL_FIRST}:E{PL_LAST}", STAGES, "Stage — pick from the nine")
dv_list(ws, f"C{PL_FIRST}:C{PL_LAST}", ["pump dealer", "bulk diesel", "lubricant distributor"])
dv_list(ws, f"D{PL_FIRST}:D{PL_LAST}",
        ["referral", "OMC TM", "association", "CA", "field visit", "WhatsApp", "inbound"],
        "Source channel — mandatory")
dv_list(ws, f"M{PL_FIRST}:M{PL_LAST}", ["₹599", "₹1,799", "₹4,999", "Enterprise", "—"])
dv_list(ws, f"O{PL_FIRST}:O{PL_LAST}", ["Y", "N"])
dv_list(ws, f"Q{PL_FIRST}:Q{PL_LAST}", ["CEO", "Domain dir.", "Third dir.", "—"])
ws.conditional_formatting.add(f"G{PL_FIRST}:G{PL_LAST}", CellIsRule(
    operator="greaterThan", formula=["14"], fill=RED_FILL))
ws.conditional_formatting.add(f"I{PL_FIRST}:I{PL_LAST}", CellIsRule(
    operator="lessThan", formula=["TODAY()"], fill=AMB_FILL))
SR = f"$E${PL_FIRST}:$E${PL_LAST}"


def reached(i):
    """Rows that reached stage i or any stage past it (Lost excluded — see the note)."""
    return "+".join(f'COUNTIF({SR},"{s}")' for s in STAGES[i:8])


r = PL_LAST + 2
band(ws, r, "B. Stages and exit tests — a stage moves only on its checkable event", 17); r += 1
header(ws, r, ["Stage", "Exit test — the checkable event", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Rows here now"]); r += 1
ST_FIRST = r
exits = [
 ("Target", "Named pump, owner, phone, source, ICP score in the Target List tab"),
 ("Contacted", "A TWO-WAY exchange. A broadcast is not contact"),
 ("Visited", "You stood at his pump and the OWNER was there"),
 ("Demoed", "Ten minutes on HIS customers and HIS rupee figures, decision-maker present (M06)"),
 ("Trial started", "Tenant created, balances entered, start and day-14 dates in his thread"),
 ("Paying", "Money received against a VSYST GST invoice"),
 ("Activated (both sides)", "≥1 customer firm linked AND ≥1 invoice within 14 days"),
 ("Referring", "A NAMED introduction that reached a two-way conversation"),
 ("Lost", "A no, or a next date blank for 30 days. Reason recorded verbatim"),
]
for s, t in exits:
    put(ws, r, [s, t] + [None] * 14 + [f'=COUNTIF({SR},"{s}")'], example=True, formula_cols=(17,), fmt={17: "0"})
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=16)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=17).font = Font(sz=10, color=INK)
    r += 1
r += 1
band(ws, r, "C. Conversion math — computed off your own rows, then argued with", 17); r += 1
header(ws, r, ["Ratio", "Reached the earlier stage", "Reached the later stage", "Rate", "Benchmark", "Reading",
               "", "", "", "", "", "", "", "", "", "", ""]); r += 1
CM_FIRST = r
ratios = [
 ("Demo → trial", 3, 4, "40–50% (lesson 06 §1)", "Below it, the demo is not landing on his rupee figures."),
 ("Demo → paying", 3, 5, "~32% SMB (Optifai)", "The headline number. Read it only past 20 demos."),
 ("Trial → paying", 4, 5, "35–45%; beat the unassisted 18.2% (First Page Sage)", "A trial nobody visits on day 14 converts like an unassisted one."),
 ("Paying → activated", 5, 6, "60–75% — the rung marketing owns alone (HC10)", "Paying is not activated. Both sides live, or it does not count."),
]
for name, a, b_, bench, read in ratios:
    put(ws, r, [name, f"={reached(a)}", f"={reached(b_)}",
                f"=IF(B{r}=0,\"\",C{r}/B{r})", bench, read],
        formula_cols=(2, 3, 4), fmt={2: "0", 3: "0", 4: PCT})
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=17)
    ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
TOUCH_ROW = r
put(ws, r, ["Touches per win", "", f"={reached(5)}",
            f"=IF(C{r}=0,\"\",B{r}/C{r})", "five (Cirrus Insight)",
            "Type the follow-ups you logged this quarter into B. If it is under three, you are quitting too early."],
    formula_cols=(3, 4), input_cols=(2,), fmt={2: "0", 3: "0", 4: "0.0"})
ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=17)
ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
r += 2
band(ws, r, "D. The Monday review — ten minutes, in this order", 17); r += 1
header(ws, r, ["#", "Step", "Count now", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
mon = [
 ("Stalled rows first — last touch over 14 days. Each gets a next action with a date, or Lost. Nothing leaves 'still thinking'.",
  f'=SUMPRODUCT((A{PL_FIRST}:A{PL_LAST}<>"")*(E{PL_FIRST}:E{PL_LAST}<>"Lost")*(G{PL_FIRST}:G{PL_LAST}<>"")*(G{PL_FIRST}:G{PL_LAST}>14))'),
 ("Blank next actions — fill or kill.",
  f'=SUMPRODUCT((A{PL_FIRST}:A{PL_LAST}<>"")*(E{PL_FIRST}:E{PL_LAST}<>"Lost")*(I{PL_FIRST}:I{PL_LAST}=""))'),
 ("Next dates already past — a date you have walked past is a blank date.",
  f'=SUMPRODUCT((I{PL_FIRST}:I{PL_LAST}<>"")*(I{PL_FIRST}:I{PL_LAST}<TODAY())*(E{PL_FIRST}:E{PL_LAST}<>"Lost"))'),
 ("Stage moves since last Monday, read aloud against the exit test (section B).", None),
 ("The five counts to the scorecard: visits, demos, trials, paying, activated (M13 → T05).", None),
 ("Conversion math (section C) — and after 20 demos, replace every benchmark with your own.", None),
]
for i, (step, f) in enumerate(mon, start=1):
    put(ws, r, [i, step, f], formula_cols=(3,) if f else (), fmt={3: "0"})
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28
    r += 1
r += 1
band(ws, r, "E. ERPNext CRM mapping — the sheet dies the day this goes live. No parallel spreadsheet.", 17); r += 1
header(ws, r, ["Sheet stage", "ERPNext", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
for a, b_ in [("Target, Contacted", "Lead, walking Lead → Open → Replied"),
              ("Visited, Demoed, Trial started", "Opportunity, raised against the Lead"),
              ("Paying", "Customer + Sales Invoice + Payment Entry; the Lead reads Converted"),
              ("Activated (both sides), Referring", "No equivalent — marketing-only columns, which is why this sheet exists")]:
    put(ws, r, [a, b_], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=17)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
note(ws, r,
     "Counting rule for section C: a row counts as having REACHED a stage if it currently sits at that stage or any later one — that is what makes a snapshot readable. "
     "A row marked Lost no longer carries the stage it reached, so the reached-counts understate slightly; write the furthest stage into the objection cell when you mark a row Lost. "
     "40–60% of an average B2B pipeline ends in 'no decision' rather than a loss (Dixon, via 6sense) — the stalled-row count in section D is the only defence against it.", 17)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── Marketing Scorecard
ws = sheet("Marketing Scorecard", "Marketing Scorecard — the eight weekly numbers (HC8), thirteen weeks trailing",
  "Fills M13 §A. Eight rows, no more. One owner enters each number personally by Monday 09:30; the 09:40 huddle reads the sheet, never builds it. "
  "No metric without a source. Fill weeks left to right without gaps (Last = last filled week). Percentages as fractions (0.4 = 40%). "
  "Red two weeks running escalates automatically onto the COO's issues list at the 10:00 ops meeting. Goals and red lines are illustrative — VERIFY LIVE at quarterly planning.",
  [4, 44, 14, 30, 9, 15, 13, 12] + [9] * 13 + [12, 9, 11, 34])
ws["H3"] = "Week starting →"; ws["H3"].font = Font(sz=9, b=True, color=MUTED)
ws["I3"] = date(2026, 8, 31); ws["I3"].number_format = DT; ws["I3"].fill = IN_FILL; ws["I3"].border = BOX
for i in range(1, 13):
    c = ws.cell(row=3, column=9 + i, value=f"={get_column_letter(9 + i - 1)}3+7")
    c.number_format = DT; c.fill = FM_FILL; c.border = BOX
header(ws, 4, ["#", "Metric (HC8)", "Owner", "Source", "Unit", "Direction", "Goal (Green)", "Red line"]
       + [f"W{i}" for i in range(1, 14)] + ["Last", "Trend", "Status", "Notes"])
ws.freeze_panes = "C5"
SC_FIRST = 5
metrics = [
 ("Pump visits", "Domain dir.", "Pipeline tab — visits logged the same evening", "count", "Higher is better", 8, 4,
  [9, 7, 12], ""),
 ("Demos done", "Domain dir.", "Pipeline tab, stage Demoed", "count", "Higher is better", 5, 2,
  [2, 3, 4], "IS T05 row 2 — entered once here, read there."),
 ("Trials started", "Third dir.", "Mongo dealer_msts", "count", "Higher is better", 3, 1,
  [1, 0, 2], ""),
 ("Dealers activated (≥1 customer linked + ≥1 invoice, 14 days)", "Third dir.", "Saved Mongo aggregation", "count", "Higher is better", 2, 0,
  [None, None, None], "Blank on purpose: aggregation not saved yet — CEO, this week. A blank with a reason beats a filled cell with a guess. IS T05 row 3."),
 ("Customers activated per activated dealer", "Third dir.", "dealer_custs × rate_msts", "ratio", "Higher is better", 3, 1,
  [None, None, None], "Same debt as row 4. Feeds T05 rows 4 and 5."),
 ("Referrals received", "Domain dir.", "Referral Tracker tab, M09", "count", "Higher is better", 2, 0,
  [0, 1, 1], "M13 states this row as 'asked / received', goal 5 / 2. One cell cannot RAG two numbers, so the scored number is RECEIVED; asks are counted on the Referral Tracker (goal 5, red 2) and read alongside."),
 ("WhatsApp reply rate", "Third dir.", "WhatsApp Business → Stats; replies within 48h ÷ recipients", "%", "Higher is better", 0.4, 0.2,
  [0.31, 0.28, 0.4], "Computed on the Content Calendar tab."),
 ("CAC by channel — best channel, rolling 90 days", "CEO", "CAC & Payback tab", "₹", "Lower is better", 8000, 20000,
  [None, None, None], "The only lagging row; the seven above exist to move it. Blank until a channel has ≥3 paying dealers."),
]
r = SC_FIRST
for i, (m, own, src, unit, direc, goal, red, vals, nt) in enumerate(metrics, start=1):
    f = RS if unit == "₹" else (PCT if unit == "%" else ("0.0" if unit == "ratio" else "#,##0"))
    put(ws, r, [i, m, own, src, unit, direc, goal, red], fmt={7: f, 8: f})
    for cc in range(9, 22):
        cell = ws.cell(row=r, column=cc)
        cell.number_format = f; cell.border = BOX; cell.fill = IN_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for j, v in enumerate(vals):
        if v is None: continue
        cell = ws.cell(row=r, column=9 + j, value=v)
        cell.font = Font(i=True, color="808080"); cell.fill = EX_FILL
    last = ws.cell(row=r, column=22, value=f'=IF(COUNTA(I{r}:U{r})=0,"",INDEX(I{r}:U{r},1,COUNTA(I{r}:U{r})))')
    last.number_format = f; last.fill = FM_FILL; last.border = BOX
    trend = ws.cell(row=r, column=23, value=(
        f'=IF(COUNTA(I{r}:U{r})<2,"",IF(V{r}>INDEX(I{r}:U{r},1,COUNTA(I{r}:U{r})-1),"▲",'
        f'IF(V{r}<INDEX(I{r}:U{r},1,COUNTA(I{r}:U{r})-1),"▼","►")))'))
    trend.fill = FM_FILL; trend.border = BOX
    trend.alignment = Alignment(horizontal="center")
    st = ws.cell(row=r, column=24, value=(
        f'=IF(V{r}="","",IF(F{r}="Higher is better",IF(V{r}>=G{r},"Green",IF(V{r}<=H{r},"Red","Amber")),'
        f'IF(V{r}<=G{r},"Green",IF(V{r}>=H{r},"Red","Amber"))))'))
    st.fill = FM_FILL; st.border = BOX; st.font = Font(b=True)
    st.alignment = Alignment(horizontal="center")
    nc = ws.cell(row=r, column=25, value=nt)
    nc.border = BOX; nc.fill = IN_FILL
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    nc.font = Font(sz=9, color=MUTED)
    ws.row_dimensions[r].height = 34
    r += 1
SC_LAST = r - 1
dv_list(ws, f"F{SC_FIRST}:F{SC_LAST}", ["Higher is better", "Lower is better"])
dv_list(ws, f"C{SC_FIRST}:C{SC_LAST}", ["CEO", "Domain dir.", "Third dir."])
dv_list(ws, f"E{SC_FIRST}:E{SC_LAST}", ["count", "%", "₹", "ratio", "days", "hours"])
for word, fill in [("Red", RED_FILL), ("Amber", AMB_FILL), ("Green", GRN_FILL)]:
    ws.conditional_formatting.add(f"X{SC_FIRST}:X{SC_LAST}",
        CellIsRule(operator="equal", formula=[f'"{word}"'], fill=fill))
r += 1
note(ws, r,
     "Rules that make this sheet worth filling: eight rows, no more (HC8) — a ninth number means one of these eight is not earning its place. "
     "History is never overwritten; a wrong number is corrected in place with a note, not erased. On track / off track only at the huddle — the argument belongs at the 10:00 ops meeting. "
     "Rows 1–7 are controllable inputs you can move this week; row 8 is the lagging number they exist for. "
     "If a marketing number and a COO number disagree, CEO lesson 17 §6.6 wins and this sheet is wrong.", 25)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── CAC & Payback
ws = sheet("CAC & Payback", "CAC & Payback — what each channel cost per paying dealer, and how long it takes to earn back",
  "Fills M13 §B–D. One row per channel per quarter; never a row called 'marketing'. Founder hours are time-boxed and logged the same day — leave them out and the cheapest-looking channel is the one that ate 68 hours. "
  "Travel is kept separate so channels compare. A CAC is real at ≥3 paying dealers, not before. Every figure below is illustrative — VERIFY LIVE.",
  [30, 16, 14, 17, 15, 17, 15, 14, 14, 15, 22, 40])
band(ws, 4, "A. House inputs — set once a year, never mid-quarter", 12)
header(ws, 5, ["Input", "Value", "Unit", "", "", "", "", "", "", "", "", "How to set it honestly"])
ws.freeze_panes = "A6"
RATE = 6
put(ws, 6, ["Imputed founder rate", 400, "₹/hour", None, None, None, None, None, None, None, None,
    "(₹30,000 salary + ~₹8,000 travel/phone for the field associate we would hire instead) ÷ 22 days ÷ 8 h ≈ ₹216, DOUBLED because the founder also demos, prices and closes. House rule ₹400. VERIFY LIVE against Raipur associate comp (HC1). Move it at quarterly planning only, or CAC moves for reasons unrelated to marketing."],
    example=True, fmt={2: RS}, input_cols=(2,))
ws.row_dimensions[6].height = 46
MRR_R, GST_R, ARPA_R, GM_R, GP_R = 7, 8, 9, 10, 11
put(ws, 7, ["MRR this month", 10500, "₹", None, None, None, None, None, None, None, None,
    "Billed and collected, ex-GST. Not signed, not promised."], example=True, fmt={2: RS}, input_cols=(2,))
put(ws, 8, ["Paying GSTINs", 7, "count", None, None, None, None, None, None, None, None,
    "The billing unit. Users, staff and drivers are unlimited and free."], example=True, fmt={2: "#,##0"}, input_cols=(2,))
put(ws, 9, ["ARPA (₹/month per paying GSTIN)", f"=IF(B{GST_R}=0,\"\",B{MRR_R}/B{GST_R})", "₹", None, None, None, None, None, None, None, None,
    "≈ ₹1,500 ex-GST on a mid-skewed mix of ₹599 / ₹1,799 / ₹4,999 + Enterprise — VERIFY LIVE, owner sign-off pending."],
    formula_cols=(2,), fmt={2: RS})
put(ws, 10, ["Gross margin", 0.80, "%", None, None, None, None, None, None, None, None,
    "(revenue − cost to serve) ÷ revenue. Cost to serve is OTP, hosting and support. 80% is the working assumption; CEO-Docs C12 owns the real number."],
    example=True, fmt={2: PCT}, input_cols=(2,))
put(ws, 11, ["Gross profit per dealer per month", f"=IF(B{ARPA_R}=\"\",\"\",B{ARPA_R}*B{GM_R})", "₹", None, None, None, None, None, None, None, None,
    "≈ ₹1,200. This is what actually pays back the acquisition — never use revenue."], formula_cols=(2,), fmt={2: RS})
r = 13
band(ws, r, "B. Spend by channel, this quarter — and what each dealer from it cost", 12); r += 1
header(ws, r, ["Channel", "Cash spend ₹", "Founder hours", "Imputed founder cost ₹", "Travel & fuel ₹",
               "Total channel cost ₹", "New paying dealers", "CAC ₹", "Free months given", "Payback (months)",
               "Real yet? (≥3 dealers)", "Note"]); r += 1
CH_FIRST = r
channels = [
 ("Referral (Dealer Dost)", 6000, 24, 1250, 3, 2, True,
  "₹6,000 of bounties, 24 h of asks and follow-ups, ₹1,250 of travel. Two free months given, which cut gross profit — so they land on payback, not on CAC."),
 ("OMC intro (IOCL territory manager)", 2000, 68, 3700, 2, 0, True,
  "Almost entirely founder time. The relationship is real; the hours are the cost nobody writes down."),
 ("Field visits (cold, in-cluster)", 1000, 62, 5600, 2, 0, True,
  "The honest floor for walking in. Compare it against referral before you plan another week of it."),
 ("Association / Dealer Day", "", "", "", "", "", False, ""),
 ("CA / tax consultant", "", "", "", "", "", False, ""),
 ("WhatsApp / content", "", "", "", "", "", False, ""),
 ("Inbound (site, Play Store, GBP)", "", "", "", "", "", False, ""),
]
for name, cash, hours, travel, dealers, free, ex, nt in channels:
    put(ws, r, [name, cash, hours, f"=IF(C{r}=\"\",\"\",C{r}*$B${RATE})", travel,
                f"=IF(COUNT(B{r},D{r},E{r})=0,\"\",N(B{r})+N(D{r})+N(E{r}))",
                dealers,
                f"=IF(OR(F{r}=\"\",G{r}=\"\",G{r}=0),\"\",F{r}/G{r})",
                free,
                f"=IF(OR(H{r}=\"\",$B${GP_R}=\"\",$B${GP_R}=0),\"\",H{r}/$B${GP_R}+N(I{r}))",
                f"=IF(G{r}=\"\",\"\",IF(G{r}>=3,\"Yes — decide on it\",\"No — not a CAC yet\"))",
                nt],
        example=ex, fmt={2: RS, 3: "0", 4: RS, 5: RS, 6: RS, 7: "0", 8: RS, 9: "0", 10: "0.0"},
        formula_cols=(4, 6, 8, 10, 11), input_cols=(1, 2, 3, 5, 7, 9, 12))
    ws.row_dimensions[r].height = 32
    r += 1
CH_LAST = r - 1
BLEND = r
put(ws, r, ["BLENDED — sanity check only, never a decision",
            f"=SUM(B{CH_FIRST}:B{CH_LAST})", f"=SUM(C{CH_FIRST}:C{CH_LAST})",
            f"=SUM(D{CH_FIRST}:D{CH_LAST})", f"=SUM(E{CH_FIRST}:E{CH_LAST})",
            f"=SUM(F{CH_FIRST}:F{CH_LAST})", f"=SUM(G{CH_FIRST}:G{CH_LAST})",
            f"=IF(G{r}=0,\"\",F{r}/G{r})", "",
            f"=IF(OR(H{r}=\"\",$B${GP_R}=\"\",$B${GP_R}=0),\"\",H{r}/$B${GP_R})", "",
            "Blended CAC moves when the MIX moves — a quarter with more referrals looks like a quarter where you got better at marketing. It cannot tell you which channel to feed."],
    formula_cols=(2, 3, 4, 5, 6, 7, 8, 10),
    fmt={2: RS, 3: "0", 4: RS, 5: RS, 6: RS, 7: "0", 8: RS, 10: "0.0"})
for c in range(1, 13):
    ws.cell(row=r, column=c).font = Font(b=True, color=INK)
ws.row_dimensions[r].height = 32
ws.conditional_formatting.add(f"J{CH_FIRST}:J{CH_LAST}", CellIsRule(
    operator="greaterThan", formula=["12"], fill=RED_FILL))
r += 2
band(ws, r, "C. Monthly channel review — first Monday, 45 minutes, one verdict per channel", 12); r += 1
header(ws, r, ["Channel", "CAC ₹", "× blended", "Paying dealers", "Payback (months)", "Suggested verdict",
               "Decided verdict", "Reason — required, written in the cell", "", "", "", ""]); r += 1
RV_FIRST = r
for i in range(len(channels)):
    src = CH_FIRST + i
    put(ws, r, [f"=A{src}", f"=H{src}",
                f"=IF(OR(H{src}=\"\",$H${BLEND}=\"\",$H${BLEND}=0),\"\",H{src}/$H${BLEND})",
                f"=G{src}", f"=J{src}",
                f'=IF(OR(H{src}="",G{src}=0),"no data yet",'
                f'IF(H{src}>2*$H${BLEND},"KILL candidate",'
                f'IF(AND(H{src}<=$H${BLEND},G{src}>=3),"SCALE candidate","KEEP")))',
                "", ""],
        formula_cols=(1, 2, 3, 4, 5, 6), input_cols=(7, 8),
        fmt={2: RS, 3: "0.00", 4: "0", 5: "0.0"})
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
    ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
RV_LAST = r - 1
dv_list(ws, f"G{RV_FIRST}:G{RV_LAST}", ["kill", "keep", "scale"], "Verdict")
r += 1
header(ws, r, ["The verdict rules — a verdict without a rule is a mood", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
for a, b_ in [("KILL", "Two months above 2× blended CAC with no improving trend. Write the reason in the cell: a killed channel with a reason can be revisited; one that quietly stopped cannot."),
              ("KEEP", "At parity with blended. Nothing changes; it stays on the sheet."),
              ("SCALE", "Two consecutive reviews beating blended AND ≥3 paying dealers from it. Raise capped at +50% of that line, one channel at a time. HC4's order still binds — no paid ads before ~50 dealers and a landing page with real proof.")]:
    put(ws, r, [a, b_], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
note(ws, r,
     "Reading the worked example: referral cost ₹16,850 for 3 dealers — CAC ₹5,617, payback 4.7 months, 6.7 after the two free months. "
     "The IOCL introduction cost ₹32,900 for 2 — CAC ₹16,450, payback 13.7. Cold field visits cost ₹31,400 for 2 — ₹15,700, payback 13.1. "
     "Blended is ₹81,150 ÷ 7 = ₹11,593, which flatters the expensive channels and punishes the cheap one. That is exactly why the decision is taken per channel. "
     "Bar for payback: CAC recovered under 12 months. Free months given are added to payback, never to CAC — they cut gross profit, not acquisition cost.", 12)
ws.row_dimensions[r].height = 60

# ─────────────────────────────────────────────── Campaign Budget
ws = sheet("Campaign Budget", "Campaign Budget — one campaign on one page: brief, budget, weekly log, post-mortem, districts",
  "Fills M12. One brief per campaign, signed by all three founders before a rupee is spent. ONE number, X to Y by a date, and queryable. "
  "Founder time (~15 h/week × 3) sits OUTSIDE the cash budget and is stated so nobody calls the campaign free. Cash inside ₹30–50k/month while pre-revenue (HC7). "
  "Filled here with the Raipur 100 example — every figure illustrative.",
  [30, 18, 16, 16, 13, 18, 12, 44, 30, 22])
band(ws, 4, "A. The one-page brief", 10)
header(ws, 5, ["Field", "Fill", "", "", "", "", "", "", "", ""])
ws.freeze_panes = "A6"
r = 6
brief = [
 ("Campaign name · ladder rung", "Raipur 100 · rung 2 (the founding ten → the district)"),
 ("Goal + THE one number", "Activated dealers 1 → 10 by the end of Week 12. Activated = ≥3 customer firms linked AND ≥1 invoice within 14 days. Query: CEO's saved Mongo aggregation over dealer_custs + invoices, run every Monday 09:00."),
 ("Audience · list source", "The 100 named pumps within 60 km on the Target List tab, plus the credit customers each signed dealer names at onboarding."),
 ("Message", "\"पिछले महीने रेट को लेकर कितनी बहस हुई?\" — fear of loss, a rupee number, and the dealer next door (HC5, M04)."),
 ("Offer", "14-day no-card trial, on-site setup, ledger migration, founding-dealer listing. Price VERIFY LIVE — owner sign-off pending."),
 ("Channels (1–2, in HC4 order)", "Referral → IOCL territory manager (Amey) → association → field visits. No paid ads on the dealer side before ~50 dealers."),
 ("Budget", "₹40,000/month × ~2.5 months ≈ ₹1.0 L cash. Founder time excluded and named: ~45 h/week across three founders ≈ ₹78,000/month imputed."),
 ("Calendar by week · owners", "One row per week in section C; a founder's name on every row and every budget line."),
 ("Risks · review cadence", "(1) Referrals dry up before the founding ten — watch asks per activated dealer. (2) The IOCL relationship stays social — watch written next steps. (3) Trials stall at day 14 — watch demo → trial. Cadence: Mon 09:30 numbers, Fri 18:00 learnings, mid-campaign kill/double, post-mortem within 5 days."),
 ("Entry test passed?", "YES — dealer #1 both sides live, ≥1 invoice a week for four weeks. Evidence: his month-end, 31 Jul 2026. Without a measured pass of the rung below, nobody signs."),
]
for f, v in brief:
    put(ws, r, [f, v], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 32
    r += 1
r += 1
band(ws, r, "B. The budget — every line carries a rupee figure and one name", 10); r += 1
header(ws, r, ["Line", "₹ planned / month", "₹ actual / month", "Variance ₹", "Variance %", "Owner", "Cut order", "What it buys", "", ""]); r += 1
BG_FIRST = r
budget = [
 ("Fuel + travel (~8 visits/week, 60 km radius)", 10000, 11500, "Domain director", "7th", "~60 pump visits within 60 km"),
 ("Printing — leave-behind + GST/TCS booklet", 6000, 5200, "Third director", "5th", "~150 leaflets + 30 booklets"),
 ("Dealer Day / association (chai + CA session, 25–40 seats)", 12000, 12000, "Domain director", "3rd", "A ₹24k event per quarter, amortised"),
 ("WhatsApp Business — BSP + templates, priced per DELIVERED message", 4000, 3600, "Third director", "4th", "~1,500 utility + ~500 marketing messages. ≈₹0.86 marketing / ₹0.115 utility + BSP markup + 18% GST — VERIFY LIVE"),
 ("Tools / software (CRM, design, landing page)", 0, 0, "Third director", "2nd", "Free tiers until they break"),
 ("Referral payouts — Dealer Dost bounty accrual (2 × ₹2,000)", 4000, 4000, "Third director", "8th — never cut", "The cheapest channel we have. Cut this last, or not at all"),
 ("Testimonial video (phone-shot + edit)", 2000, 1500, "CEO", "6th", "One shoot, phone + freelance edit"),
 ("Contingency / channel experiment", 2000, 3000, "Third director", "1st — cut first", "One Channel Experiments test"),
]
for line, plan, act, own, cut, buys in budget:
    put(ws, r, [line, plan, act, f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",C{r}-B{r})",
                f"=IF(OR(B{r}=\"\",B{r}=0,C{r}=\"\"),\"\",(C{r}-B{r})/B{r})", own, cut, buys],
        example=True, fmt={2: RS, 3: RS, 4: RS, 5: PCT}, formula_cols=(4, 5), input_cols=(2, 3))
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
    ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28
    r += 1
BG_LAST = r - 1
put(ws, r, ["TOTAL / month — must stay inside the ₹30–50k band (HC7)",
            f"=SUM(B{BG_FIRST}:B{BG_LAST})", f"=SUM(C{BG_FIRST}:C{BG_LAST})",
            f"=C{r}-B{r}", f"=IF(B{r}=0,\"\",(C{r}-B{r})/B{r})", "", "",
            "Any single line moving by more than ₹5,000 needs a dated CEO sign-off and a line in the decision log. The envelope itself is a runway call."],
    formula_cols=(2, 3, 4, 5), fmt={2: RS, 3: RS, 4: RS, 5: PCT})
for c in range(1, 11):
    ws.cell(row=r, column=c).font = Font(b=True, color=INK)
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
BG_TOT = r
r += 1
put(ws, r, ["Founder time — NOT cash, never cut, always stated", "≈₹78,000", "", "", "", "All three", "—",
            "45 h/week × ₹400 imputed. Roughly twice the cash budget: saving ₹5,000 of cash by spending ten extra founder hours is a losing trade."],
    example=True)
ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
ws.conditional_formatting.add(f"D{BG_FIRST}:D{BG_TOT}", CellIsRule(
    operator="greaterThan", formula=["0"], fill=RED_FILL))
r += 2
band(ws, r, "C. The weekly campaign log — filled Monday 09:30, zeros included", 10); r += 1
header(ws, r, ["Week", "Visits", "Demos", "Trials", "Activations", "Referrals", "Spend ₹",
               "Learning (one line)", "Change made Monday", ""]); r += 1
LG_FIRST = r
for w in range(1, 13):
    if w == 6:
        put(ws, r, ["W6", 9, 5, 2, 0, 1, 9400,
                    "Both cold walk-ins refused the demo; both introduced pumps sat through all ten minutes.",
                    "Two cold-visit slots become referral asks."], example=True, fmt={7: RS})
    else:
        put(ws, r, [f"W{w}", "", "", "", "", "", "", "", ""], fmt={7: RS},
            input_cols=(2, 3, 4, 5, 6, 7, 8, 9))
    ws.row_dimensions[r].height = 26
    r += 1
LG_LAST = r - 1
put(ws, r, ["TOTAL"] + [f"=SUM({get_column_letter(c)}{LG_FIRST}:{get_column_letter(c)}{LG_LAST})" for c in range(2, 8)]
    + ["Skipping the zero weeks removes the diagnosis. Back-filling on Friday invents one.", ""],
    formula_cols=(2, 3, 4, 5, 6, 7), fmt={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "#,##0", 7: RS})
for c in range(1, 11):
    ws.cell(row=r, column=c).font = Font(b=True, color=INK)
LG_TOT = r
r += 2
band(ws, r, "D. The post-mortem — six boxes, 45 minutes, blameless, within five days of the close", 10); r += 1
header(ws, r, ["Box", "Fill", "", "", "", "", "", "", "", ""]); r += 1
pm = [
 ("Hypothesis", "A named list worked warm — referral, then OMC introduction, then association — converts at about 10% of the list into activated dealers in 12 weeks."),
 ("Result vs the number", "8 activated against a target of 10. The gap opened at demo → trial, not at the door: getting in was fine, getting the tenant created was not."),
 ("CAC by channel", "See the table below — one row per channel, spend ÷ activated. The same rows go to the Channel Experiments tab."),
 ("Keep", "Referral asks at the activation moment; the CA-led session as a warm-up; the same-evening pipeline update."),
 ("Kill", "Cold field visits — ₹42,000 freed, nothing to show. Stops Monday."),
 ("Change", "One change only: the day-14 trial meeting becomes a booked calendar entry at trial start, owned by the domain director."),
 ("Next campaign · date", "Rung 3, entry test as numbers: ≥10 paying dealers, one channel producing ≥3 repeatably, CAC known ±30%, activation ≥60%. Brief written by 15 Dec 2026."),
]
for b_, v in pm:
    put(ws, r, [b_, v], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
header(ws, r, ["Channel", "Spend ₹", "Activated", "CAC ₹", "Verdict", "Why", "", "", "", ""]); r += 1
PM_FIRST = r
pmc = [("Referral", 22000, 4, "Keep", "Double the ask rate — it is the cheapest thing on this sheet"),
       ("IOCL territory manager", 8000, 2, "Keep", "One meeting a month, no more; the cost is founder hours, not cash"),
       ("Association / Dealer Day", 30000, 2, "Change", "One CA-led session, not two"),
       ("Cold field visits", 42000, 0, "Kill", "₹42,000 freed. Revisit only with a written reason")]
for name, spend, act, verd, why in pmc:
    put(ws, r, [name, spend, act, f"=IF(OR(C{r}=\"\",C{r}=0),\"—\",B{r}/C{r})", verd, why],
        example=True, fmt={2: RS, 3: "0", 4: RS}, formula_cols=(4,))
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
    ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
PM_LAST = r - 1
put(ws, r, ["BLENDED", f"=SUM(B{PM_FIRST}:B{PM_LAST})", f"=SUM(C{PM_FIRST}:C{PM_LAST})",
            f"=IF(C{r}=0,\"—\",B{r}/C{r})", "",
            "Planned ₹10,000. Actual ₹12,750 — a 27.5% miss that is entirely explained by the killed row."],
    formula_cols=(2, 3, 4), fmt={2: RS, 3: "0", 4: RS})
for c in range(1, 11):
    ws.cell(row=r, column=c).font = Font(b=True, color=INK)
ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
r += 2
band(ws, r, "E. The multi-district variant — only after the rung-3 entry test passes. Before that, mark it DRAFT — not funded.", 10); r += 1
header(ws, r, ["District", "Associate", "Association / OMC letter", "Dealer Day date", "List size",
               "The number (90 days)", "Activated at T+30", "T+30 gate", "Owner", "Note"]); r += 1
MD_FIRST = r
districts = [("Raipur", "—  (founders)", "sent, on letterhead", date(2026, 9, 15), 100, 15, 6, "Domain dir.",
              "The proven district. Everything below copies its playbook or it does not run."),
             ("Durg–Bhilai", "", "", "", "", 15, "", "", ""),
             ("Bilaspur", "", "", "", "", 15, "", "", ""),
             ("Rajnandgaon", "", "", "", "", 15, "", "", "")]
for i, (d, a, letter, day, size, num, act, own, nt) in enumerate(districts):
    ex = (i == 0)
    put(ws, r, [d, a, letter, day, size, num, act,
                f'=IF(G{r}="","",IF(G{r}>=5,"GO",IF(G{r}>=3,"WATCH","STOP")))', own, nt],
        example=ex, fmt={4: DT, 5: "#,##0", 6: "#,##0", 7: "#,##0"}, formula_cols=(8,),
        input_cols=(1, 2, 3, 4, 5, 6, 7, 9, 10))
    ws.row_dimensions[r].height = 28
    r += 1
MD_LAST = r - 1
dv_list(ws, f"C{MD_FIRST}:C{MD_LAST}", ["not sent", "sent", "sent, on letterhead"])
for word, fill in [("STOP", RED_FILL), ("WATCH", AMB_FILL), ("GO", GRN_FILL)]:
    ws.conditional_formatting.add(f"H{MD_FIRST}:H{MD_LAST}",
        CellIsRule(operator="equal", formula=[f'"{word}"'], fill=fill))
r += 1
note(ws, r,
     "Rules for section E: one row per district, the SAME number definition everywhere, one named owner each. Three coordination artefacts, no more — a brief per district, one scorecard with a district row plus a company total, and a daily 15-minute stand-up. "
     "Size districts from PPAC's state-wise retail-outlet file — VERIFY LIVE. The budget shape flips to ~65–70% people, roughly ₹5.5–8 L over 90 days: it breaks HC7's band deliberately, funded from revenue or a round, never from runway. "
     "Run before CAC is known, this is four times the burn with no diagnosis.", 10)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── Content Calendar
ws = sheet("Content Calendar", "Content Calendar — one piece a week, and whether it earned a reply",
  "Fills M11 §D. One row, one piece, forty minutes of editing: four hours means the format is wrong. Row filled Friday, broadcast Tuesday. "
  "Never Sunday; 10:30–12:30 or 16:00–18:00, outside the pump's rush. One broadcast per list per week, maximum — and only to dealers who opted in AND saved your number. "
  "Reply rate here is scorecard row 7. Zero replies twice means the row is deleted, not debated.",
  [7, 34, 20, 16, 13, 13, 30, 9, 12, 12, 11, 12, 30, 11])
band(ws, 4, "A. The calendar", 14)
header(ws, 5, ["Week", "Piece", "Type", "Persona", "Shoot owner", "Edit owner",
               "Channels (Status → broadcast → Shorts → GBP → LinkedIn)", "Posted ✓",
               "Recipients", "Replies @48h", "Reply rate", "Views @48h", "Reuse", "Verdict"])
ws.freeze_panes = "B6"
CC_FIRST = 6
cc_ex = [
 ("W1", "\"3 din se 10 minute\" — Sharma ji", "Testimonial, 84s", "Owner, 52", "Domain dir", "Third dir",
  "Status, broadcast, Shorts, GBP", "✓", 34, 7, 260, "Still + pull-quote for the leave-behind", "keep"),
 ("W2", "\"Month-end in 10 minutes\"", "Screen recording, 110s", "Accountant", "Third dir", "Third dir",
  "Status, broadcast, Shorts", "✓", 34, 3, 190, "Three-line text for WhatsApp openers", "kill"),
 ("W3", "\"जो रेट तय हुआ था\"", "Story, 62s", "Owner + son", "Domain dir", "Third dir",
  "Status, broadcast, Shorts, GBP", "✓", 36, 9, 410, "Landing-page paragraph + proof slot 2", "keep"),
 ("W4", "\"TCS ₹50L ke baad\", with CA Bhatia", "Explainer with CA, 118s", "Son / accountant", "Third dir", "Third dir",
  "Status, broadcast, Shorts, LinkedIn", "✓", 36, 4, 150, "Warm-up asset for the CA partner playbook, not a broadcast", "kill"),
]
r = CC_FIRST
for row in cc_ex:
    put(ws, r, list(row[:10]) + [f"=IF(OR(I{r}=\"\",I{r}=0),\"\",J{r}/I{r})"] + list(row[10:]),
        example=True, fmt={9: "#,##0", 10: "#,##0", 11: PCT, 12: "#,##0"}, formula_cols=(11,))
    ws.row_dimensions[r].height = 30
    r += 1
for w in range(5, 27):
    put(ws, r, [f"W{w}", "", "", "", "", "", "", "", "", "",
                f"=IF(OR(I{r}=\"\",I{r}=0),\"\",J{r}/I{r})", "", "", ""],
        fmt={9: "#,##0", 10: "#,##0", 11: PCT, 12: "#,##0"}, formula_cols=(11,),
        input_cols=(2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14))
    r += 1
CC_LAST = r - 1
dv_list(ws, f"C{CC_FIRST}:C{CC_LAST}",
        ["Testimonial video", "Screen recording", "Story", "Explainer with CA", "Lead magnet", "Other"])
dv_list(ws, f"E{CC_FIRST}:E{CC_LAST}", ["CEO", "Domain dir", "Third dir"])
dv_list(ws, f"F{CC_FIRST}:F{CC_LAST}", ["CEO", "Domain dir", "Third dir"])
dv_list(ws, f"H{CC_FIRST}:H{CC_LAST}", ["✓", "not yet", "dropped"])
dv_list(ws, f"N{CC_FIRST}:N{CC_LAST}", ["keep", "kill"])
ws.conditional_formatting.add(f"K{CC_FIRST}:K{CC_LAST}", CellIsRule(
    operator="lessThan", formula=["0.15"], fill=RED_FILL))
r = CC_LAST + 2
band(ws, r, "B. Rows that go to the Marketing Scorecard and the campaign post-mortem", 14); r += 1
header(ws, r, ["Row", "Formula", "Value", "Goal", "Cadence", "", "", "", "", "", "", "", "", ""]); r += 1
RR = r
put(ws, r, ["WhatsApp reply rate (rolling, all filled weeks)",
            "Σ replies ÷ Σ recipients — never the average of the weekly percentages",
            f"=IF(SUM(I{CC_FIRST}:I{CC_LAST})=0,\"\",SUM(J{CC_FIRST}:J{CC_LAST})/SUM(I{CC_FIRST}:I{CC_LAST}))",
            0.15, "Weekly — one of HC8's eight, scorecard row 7"],
    formula_cols=(3,), fmt={3: PCT, 4: PCT}); r += 1
put(ws, r, ["Pieces published", "Rows marked Posted ✓",
            f"=COUNTIF(H{CC_FIRST}:H{CC_LAST},\"✓\")", 1, "Weekly, hygiene — one a week, forever"],
    formula_cols=(3,), fmt={3: "#,##0"}); r += 1
put(ws, r, ["Median views @48h", "Vanity on its own; useful only next to replies",
            f"=IF(COUNT(L{CC_FIRST}:L{CC_LAST})=0,\"\",MEDIAN(L{CC_FIRST}:L{CC_LAST}))", "", "Monthly, context only"],
    formula_cols=(3,), fmt={3: "#,##0"}); r += 1
put(ws, r, ["Click-to-chat conversion", "Chats opened ÷ page visits — hand-counted on Friday from the inbox; GA4 sees the click, only the inbox sees the conversation",
            "", 0.05, "Monthly, into the campaign post-mortem"],
    input_cols=(3,), fmt={3: PCT, 4: PCT}); r += 1
ws.conditional_formatting.add(f"C{RR}", CellIsRule(operator="lessThan", formula=["0.15"], fill=AMB_FILL))
r += 1
band(ws, r, "C. The broadcast rules — break one and the number your pipeline lives in gets banned", 14); r += 1
header(ws, r, ["Rule", "What it says", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
rules = [
 ("Opt-in", "Before any broadcast. The line must say the person agrees to receive messages AND name the business. A tick on the pump-visit form or on the site both count."),
 ("Saved your number", "A broadcast reaches only people who saved your number. They receive nothing otherwise — and you conclude, wrongly, that the channel is dead."),
 ("Cadence", "One per list per week, maximum. Never Sunday; 10:30–12:30 or 16:00–18:00."),
 ("Shape", "Proof → one line → one tap. Five lines or fewer, one link, one action."),
 ("Reply SLA", "Two working hours between 09:00 and 20:00; one working day always. A broadcast you do not staff is worse than none."),
 ("List size", "One list, ≤256 dealers. When it outgrows the app — the same message to 200+ dealers, or a second list — that is the trigger for the API, and only then."),
 ("UTMs", "utm_source, utm_medium, utm_campaign on every outbound link. House form: ?utm_source=whatsapp&utm_medium=broadcast&utm_campaign=raipur100-w07"),
 ("The CTA", "A chat, not a form. wa.me/91XXXXXXXXXX?text=<url-encoded first line> — international format, no zeroes, brackets or dashes. Nobody in this market fills a form."),
]
for a, b_ in rules:
    put(ws, r, [a, b_], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=14)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28
    r += 1
r += 1
note(ws, r,
     "Reading the four example weeks: the story-shaped pieces out-pull the explanatory ones two to one, so W5–W8 become dealer stories from other tehsils and the CA explainer survives as a partner warm-up, not a broadcast. "
     "Rolling reply rate 16.4% clears the ≥15% goal; click-to-chat 4.6% falls just short. Founder time for the month: five hours. "
     "Count replies by hand — the free WhatsApp Business app has no dashboard. India WhatsApp pricing ≈ ₹0.86 per marketing message, ₹0.115 per utility, + BSP markup ₹0.10–0.30 + 18% GST — VERIFY LIVE with the BSP.", 14)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── Referral Tracker
ws = sheet("Referral Tracker", "Referral Tracker — every Dealer Dost ask, including the ones that failed",
  "Fills M09 §C. One row per REFERRED PUMP, never 'two names given'. Logged the same day, always — even if he said no. Reviewed Friday, 30 minutes. "
  "The ask is made at the activation moment, never at the sale, and never in the first 14 days. Ask only for the two pumps he already talks to. "
  "Reward on payment, not signup — the strongest anti-fraud lever there is. All rupee figures VERIFY LIVE against the signed tier sheet.",
  [30, 32, 13, 13, 18, 14, 13, 18, 13, 34])
band(ws, 4, "A. The tracker", 10)
header(ws, 5, ["Referrer (dealer + pump)", "Referred pump (name, owner, place)", "Date asked", "Date received",
               "Stage", "Reward due ₹", "Paid on", "UTR", "Owner (ours)", "Note"])
ws.freeze_panes = "C6"
RF_FIRST = 6
REF_STAGES = ["Asked", "Name received", "Meeting agreed", "Demo", "Live", "Paying", "Declined"]
rf_ex = [
 ("Sharma ji — Dhamtari Road Fuels", "HP outlet, Tatibandh — owner N. Chandrakar", date(2026, 8, 21), date(2026, 8, 21),
  "Name received", None, None, "Domain dir.", "Three-way WhatsApp group made on the spot. Next step Monday."),
 ("Sharma ji — Dhamtari Road Fuels", "Cousin's pump, Abhanpur — owner P. Sharma", date(2026, 8, 21), date(2026, 8, 21),
  "Name received", None, None, "Domain dir.", "Forwarded our message. Meeting not yet agreed."),
 ("R. Verma — Maa Danteshwari Fuels", "Kharora highway pump — owner S. Sahu", date(2026, 7, 14), date(2026, 7, 16),
  "Live", None, None, "Domain dir.", "Went live on one outlet 12 Aug. First ₹2,000 due — CEO to sign the payout."),
 ("R. Verma — Maa Danteshwari Fuels", "Mandir Hasaud pump — owner A. Yadav", date(2026, 6, 30), date(2026, 7, 2),
  "Paying", date(2026, 8, 20), "UTR 2026082012345678", "Domain dir.", "Month-3 payment received. Full ₹5,000 paid."),
 ("M. Agrawal — Shree Balaji Filling", "Birgaon pump — owner (name not given)", date(2026, 8, 12), None,
  "Asked", None, None, "CEO", "Said he would think about it. Logged anyway — the failures are the honest half of the sheet."),
]
r = RF_FIRST
for row in rf_ex:
    put(ws, r, list(row[:5]) + [
        f'=IF(E{r}="","",IF(E{r}="Paying",5000,IF(E{r}="Live",2000,0)))'] + list(row[5:]),
        example=True, fmt={3: DT, 4: DT, 6: RS, 7: DT}, formula_cols=(6,))
    ws.row_dimensions[r].height = 28
    r += 1
for rr in range(r, r + 30):
    put(ws, rr, ["", "", "", "", "",
                 f'=IF(E{rr}="","",IF(E{rr}="Paying",5000,IF(E{rr}="Live",2000,0)))', "", "", "", ""],
        fmt={3: DT, 4: DT, 6: RS, 7: DT}, formula_cols=(6,),
        input_cols=(1, 2, 3, 4, 5, 7, 8, 9, 10))
RF_LAST = r + 29
dv_list(ws, f"E{RF_FIRST}:E{RF_LAST}", REF_STAGES, "Stage")
dv_list(ws, f"I{RF_FIRST}:I{RF_LAST}", ["CEO", "Domain dir.", "Third dir."])
ws.conditional_formatting.add(f"F{RF_FIRST}:F{RF_LAST}", CellIsRule(
    operator="greaterThan", formula=["0"], fill=GRN_FILL))
r = RF_LAST + 2
band(ws, r, "B. Totals — read every Friday, 30 minutes", 10); r += 1
header(ws, r, ["Count / amount", "Value", "Rule", "", "", "", "", "", "", ""]); r += 1
tot = [
 ("Asks made (rows logged)", f"=SUMPRODUCT(1*(A{RF_FIRST}:A{RF_LAST}<>\"\"))", "Every ask, including the ones that went nowhere. Scorecard: referrals asked, goal 5/week."),
 ("Names received", f"=SUMPRODUCT(1*(D{RF_FIRST}:D{RF_LAST}<>\"\"))", "A name is not a referral. Scorecard row 6 counts what actually arrived."),
 ("Meetings agreed", f"=COUNTIF(E{RF_FIRST}:E{RF_LAST},\"Meeting agreed\")", "This is where a name becomes a referral — the owner agreed to meet us."),
 ("Live", f"=COUNTIF(E{RF_FIRST}:E{RF_LAST},\"Live\")", "First bounty falls due here: ₹2,000 on go-live at one outlet. VERIFY LIVE."),
 ("Paying (month 3)", f"=COUNTIF(E{RF_FIRST}:E{RF_LAST},\"Paying\")", "Second bounty falls due: a further ₹3,000, ₹5,000 in total. VERIFY LIVE."),
 ("Ask → name conversion", f"=IF(SUMPRODUCT(1*(A{RF_FIRST}:A{RF_LAST}<>\"\"))=0,\"\",SUMPRODUCT(1*(D{RF_FIRST}:D{RF_LAST}<>\"\"))/SUMPRODUCT(1*(A{RF_FIRST}:A{RF_LAST}<>\"\")))",
  "Under 50% usually means the ask was generic. Never 'do you know anyone?' — ask for the two pumps he already talks to."),
 ("Total reward due ₹", f"=SUM(F{RF_FIRST}:F{RF_LAST})", "Accrue it in the campaign budget the moment a row goes Live, not when it is paid."),
 ("Total paid ₹", f"=SUMIFS(F{RF_FIRST}:F{RF_LAST},G{RF_FIRST}:G{RF_LAST},\">0\")", "Paid means a UTR is in the row. No UTR, not paid."),
 ("Outstanding ₹", None, "What we owe dealers right now. A withdrawn or delayed reward costs more trust than it bought."),
 ("Largest single referrer, ₹ this FY", f"=IF(COUNTA(A{RF_FIRST}:A{RF_LAST})=0,\"\",MAX(SUMIF(A{RF_FIRST}:A{RF_LAST},A{RF_FIRST}:A{RF_LAST},F{RF_FIRST}:F{RF_LAST})))",
  "Array formula — confirm with Ctrl+Shift+Enter in older Excel. TDS at 10% under §194R can apply once benefits to ONE recipient cross ₹20,000 in a financial year. The CA signs off before the first payout. VERIFY LIVE."),
]
TOT_FIRST = r
for n, f, rule in tot:
    put(ws, r, [n, f, rule], formula_cols=(2,) if f else (), fmt={2: RS if "₹" in n else ("0.0%" if "conversion" in n else "#,##0")})
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26
    r += 1
OUTSTANDING = TOT_FIRST + 8
ws.cell(row=OUTSTANDING, column=2).value = f"=B{TOT_FIRST + 6}-B{TOT_FIRST + 7}"
ws.cell(row=OUTSTANDING, column=2).number_format = RS
ws.cell(row=OUTSTANDING, column=2).fill = FM_FILL
ws.cell(row=OUTSTANDING, column=2).border = BOX
ws.conditional_formatting.add(f"B{TOT_FIRST + 9}", CellIsRule(
    operator="greaterThanOrEqual", formula=["20000"], fill=RED_FILL))
r += 1
band(ws, r, "C. The ask — say it, do not send it", 10); r += 1
header(ws, r, ["Moment", "What you open with", "What you ask for", "", "", "", "", "", "", ""]); r += 1
asks = [("First clean month-end", "\"Sir, is mahine month-end saaf nikla — koi ledger ka jhagda nahi.\"", "The two pumps he already talks to, by name"),
        ("First dispute avoided", "\"Wo rate wali baat OTP se ek minute mein khatam ho gayi na?\"", "The same two names"),
        ("Unprompted praise", "Stop the conversation right there", "The same two names"),
        ("NEVER", "At the sale, or in the first 14 days", "—")]
for a, b_, c_ in asks:
    put(ws, r, [a, b_, c_], example=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
note(ws, r,
     "Then remove the work: he can (1) forward our message, (2) send the name and number, or (3) make a three-way WhatsApp group and type one line. Ask for (3), accept (1). "
     "Before using any proof asset take consent on WhatsApp in his own words, and blur customer names in any screenshot. "
     "Terms, in plain words: the same PAN or GSTIN cannot sit on both sides; a name counts once the owner agrees to meet us; the reward is clawed back if the pump leaves within six months. "
     "88% of people trust recommendations from people they know, above every other marketing message — which is why this is channel one in HC4, not a nice-to-have.", 10)
ws.row_dimensions[r].height = 56

# ─────────────────────────────────────────────── Channel Experiments
ws = sheet("Channel Experiments", "Channel Experiments — one channel, two weeks, one number, a verdict",
  "Fills lesson 05 §6 and exercise 10.3; reviewed monthly per lesson 11 §7. Exactly one channel per row — two at once teach nothing about either. "
  "The kill line is written BEFORE the start date and the verdict cell stays empty until the end date, which goes in the calendar as a 15-minute meeting named 'Verdict: <channel>'. "
  "Two weeks is deliberate: long enough for an association meeting or a CA callback, short enough that a bad channel costs a fortnight.",
  [24, 46, 12, 12, 11, 13, 12, 14, 24, 22, 11, 10, 12, 13, 12, 13, 34])
band(ws, 4, "A. The experiments", 17)
header(ws, 5, ["Channel", "Hypothesis (one falsifiable sentence)", "Start", "End (+14 days)",
               "Founder hours", "Cash spend ₹", "Travel ₹", "Total cost ₹", "The number counted",
               "Kill line (written before the start)", "Contacts", "Demos", "Paying dealers", "CAC ₹",
               "Verdict", "Owner", "Notes / why"])
ws.freeze_panes = "B6"
CE_FIRST = 6
ce = [
 ("CA / tax consultant", "A Raipur CA with pump clients introduces 3+ dealers for a 10% revenue share.",
  date(2026, 9, 1), date(2026, 9, 15), 6, 2000, 800, "Qualified pump meetings booked",
  "<2 meetings in 14 days → kill", 5, 2, 1, "keep", "Third dir.",
  "Five CAs met, two introduced someone. Slow, but the introduction is warm and the accountant is the objection we hear most."),
 ("Dealer association (district)", "The district association lets us run a 30-minute GST/TCS session at a members' meeting, and 3+ owners ask for a demo.",
  date(2026, 9, 16), date(2026, 9, 30), 10, 8000, 1200, "Demo requests from attendees",
  "<3 demo requests → kill", 28, 4, 1, "keep", "Domain dir.",
  "Office-bearer's name came from the live dealer. One session per quarter, not one per month."),
 ("WhatsApp community", "Being useful in two existing dealer groups produces 2+ inbound conversations in 14 days.",
  date(2026, 10, 1), date(2026, 10, 15), 8, 0, 0, "Inbound two-way conversations",
  "<2 conversations → kill", 3, 1, 0, "kill", "Third dir.",
  "Killed with a reason: the groups are for rates and complaints, not tools. Revisit only with a dealer vouching in the group."),
 ("Hardware service engineers", "Pump automation engineers, who visit every pump quarterly, will name 5+ automated pumps for a referral fee.",
  date(2026, 10, 16), date(2026, 10, 30), 5, 1500, 900, "Named, automated pumps handed over",
  "<5 names in 14 days → kill", 2, 0, 0, "", "Domain dir.",
  "Running. Verdict meeting is in the calendar for the end date."),
]
r = CE_FIRST
for row in ce:
    put(ws, r, list(row[:7]) + [
        f"=IF(COUNT(E{r},F{r},G{r})=0,\"\",N(F{r})+N(E{r})*'CAC & Payback'!$B${RATE}+N(G{r}))"] + list(row[7:12]) + [
        f"=IF(OR(H{r}=\"\",M{r}=\"\",M{r}=0),\"—\",H{r}/M{r})"] + list(row[12:]),
        example=True, fmt={3: DT, 4: DT, 5: "0", 6: RS, 7: RS, 8: RS, 11: "0", 12: "0", 13: "0", 14: RS},
        formula_cols=(8, 14))
    ws.row_dimensions[r].height = 44
    r += 1
for rr in range(r, r + 20):
    put(ws, rr, ["", "", "", "", "", "", "",
                 f"=IF(COUNT(E{rr},F{rr},G{rr})=0,\"\",N(F{rr})+N(E{rr})*'CAC & Payback'!$B${RATE}+N(G{rr}))",
                 "", "", "", "", "",
                 f"=IF(OR(H{rr}=\"\",M{rr}=\"\",M{rr}=0),\"—\",H{rr}/M{rr})", "", "", ""],
        fmt={3: DT, 4: DT, 5: "0", 6: RS, 7: RS, 8: RS, 11: "0", 12: "0", 13: "0", 14: RS},
        formula_cols=(8, 14), input_cols=(1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13, 15, 16, 17))
CE_LAST = r + 19
dv_list(ws, f"O{CE_FIRST}:O{CE_LAST}", ["kill", "keep", "scale"], "Verdict")
dv_list(ws, f"P{CE_FIRST}:P{CE_LAST}", ["CEO", "Domain dir.", "Third dir."])
dv_list(ws, f"A{CE_FIRST}:A{CE_LAST}",
        ["Referral", "OMC TM / IOCL", "Dealer association", "CA / tax consultant", "Field visits",
         "WhatsApp community", "Hardware service engineers", "Content / SEO", "Play Store / App Store",
         "Google Business Profile", "Paid ads", "Cold email / LinkedIn", "Other"])
ws.conditional_formatting.add(f"D{CE_FIRST}:D{CE_LAST}", CellIsRule(
    operator="lessThan", formula=["TODAY()"], fill=AMB_FILL))
for word, fill in [("kill", RED_FILL), ("keep", AMB_FILL), ("scale", GRN_FILL)]:
    ws.conditional_formatting.add(f"O{CE_FIRST}:O{CE_LAST}",
        CellIsRule(operator="equal", formula=[f'"{word}"'], fill=fill))
r = CE_LAST + 2
band(ws, r, "B. The method — six fields, and the sixth is written last", 17); r += 1
header(ws, r, ["Field", "Rule", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
method = [
 ("Channel", "Exactly one. Two at once teach nothing about either."),
 ("Hypothesis", "One falsifiable sentence, with a number in it. If it cannot be wrong, it is not a hypothesis."),
 ("Effort box", "Hours and rupees, fixed up front. When the box is empty the experiment ends, whatever the result looks like."),
 ("The number", "One, and a step you control — meetings booked, names handed over. Never revenue."),
 ("Kill line", "Written BEFORE you start. Writing it afterwards is how a dead channel survives six months."),
 ("Verdict", "Kill / keep / scale, with one sentence of why. A killed channel with a reason can be revisited; one that quietly stopped cannot."),
]
for a, b_ in method:
    put(ws, r, [a, b_], example=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=17)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
band(ws, r, "C. The channel order (HC4) — and what is not open yet", 17); r += 1
header(ws, r, ["#", "Channel", "Status today", "Trigger to open it", "", "", "", "", "", "", "", "", "", "", "", "", ""]); r += 1
order = [
 (1, "Referral (Dealer Dost)", "OPEN — channel one, cheapest we have", "Already open. Ask at every activation moment."),
 (2, "OMC territory manager / IOCL relationship", "OPEN — live conversation via Amey", "Already open. One meeting a month, with a written next step each time."),
 (3, "Dealer association", "OPEN", "Already open. One session a quarter."),
 (4, "CA / tax consultant", "OPEN — under test", "Already open. See the experiment rows above."),
 (5, "Field visits", "OPEN — the honest floor", "Already open, and the most expensive thing on the sheet. Compare it against referral every month."),
 (6, "WhatsApp community", "TESTED — killed with a reason", "A live dealer willing to vouch inside the group."),
 (7, "Content / SEO", "NOT YET", "Three real dealer proofs and a landing page that converts a chat. Months, not weeks."),
 (8, "Paid ads", "NOT YET — HC4 is explicit", "~50 paying dealers AND a landing page with real proof. CEO decides, in writing."),
 (9, "Cold email / LinkedIn", "NOT YET — narrow", "Second-generation owners only. Never the 55-year-old signer."),
]
for i, ch, st, trig in order:
    put(ws, r, [i, ch, st, trig], example=True)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=17)
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
note(ws, r,
     "Total cost on this tab uses the same imputed founder rate as the CAC & Payback tab (cell B" + str(RATE) + " there), so a CAC computed here and a CAC computed there are the same number. "
     "CAC per channel flows into the CAC & Payback tab and then into scorecard row 8. "
     "This tab is the only defence against what kills bootstrapped marketing: six half-channels for six months and no way to say which one worked.", 17)
ws.row_dimensions[r].height = 46

OUT = ("/Users/shikhar/Documents/KIT/GITHUB/DZZLO_OMS/v1_79/obsidian-notes/content/"
       "vsyst-technologies/sakuradocs/CMO-Docs/toolkit/vsyst-cmo-workbook.xlsx")
wb.save(OUT)
print("OK — sheets:", wb.sheetnames)
print("Saved:", OUT)
