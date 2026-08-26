# Builds vsyst-ceo-workbook.xlsx — the CEO Toolkit workbook.
# Styling mirrors vsyst-coo-workbook.xlsx exactly.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from datetime import date

INK, MUTED, HDR_BG = "0F172A", "64748B", "334155"
EX_FILL = PatternFill("solid", fgColor="FFF9C4")   # example rows (grey italic on pale yellow)
FM_FILL = PatternFill("solid", fgColor="E8F0FE")   # computed / formula cells
SEC_FILL = PatternFill("solid", fgColor="E2E8F0")  # section band
IN_FILL  = PatternFill("solid", fgColor="DCFCE7")  # input cells you type into
HDR_FILL = PatternFill("solid", fgColor=HDR_BG)
RS = '"₹"#,##0;-"₹"#,##0'
RS2 = '"₹"#,##0.00;-"₹"#,##0.00'
PCT = '0.0%'
THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def sheet(name, title, sub, widths, tab="475569"):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    ws["A1"] = title
    ws["A1"].font = Font(sz=13, b=True, color=INK); ws["A1"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 23
    ws["A2"] = sub
    ws["A2"].font = Font(sz=9, i=True, color=MUTED); ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 26
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def band(ws, row, text, span):
    ws.cell(row=row, column=1, value=text).font = Font(sz=10, b=True, color=INK)
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SEC_FILL

def header(ws, row, cols):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(sz=10, b=True, color="FFFFFF"); c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30

def put(ws, row, vals, start=1, example=False, fmt=None, formula_cols=(), input_cols=()):
    for i, v in enumerate(vals, start=start):
        if v is None: continue
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        if example:
            c.font = Font(i=True, color="808080")
            c.fill = EX_FILL
        if i in formula_cols: c.fill = FM_FILL
        if i in input_cols and not example: c.fill = IN_FILL
        if fmt and i in fmt: c.number_format = fmt[i]
        c.alignment = Alignment(wrap_text=True, vertical="top")

# ─────────────────────────────────────────────── README
ws = sheet("README", "README — what each tab does, who owns it, and how to fill it",
  "VSYST CEO Workbook · version 1.0 · built 2026-08-25 · companion to CEO-Docs. Grey italic on pale yellow = example data, overwrite it. Pale blue = a live formula, do not type over it. Pale green = a cell you are meant to type into. Import to Google Sheets: File → Import → Upload → Replace spreadsheet, then re-check the formulas on the Cap Table and Runway tabs.",
  [5, 26, 52, 52, 30], tab=MUTED)
header(ws, 4, ["#", "Tab", "What it does", "Key formulas / how to fill", "Suggested owner · cadence"])
ws.freeze_panes = "A5"
readme = [
 ("README","This page: tab guide, colour legend, Google Sheets import note.","Read the legend in the subtitle before touching anything.","CEO · on every version bump"),
 ("Strategy Kernel","Rumelt's diagnosis / guiding policy / coherent action, plus the Playing-to-Win cascade and the assumption register that keeps the strategy honest.","Free text. The Assumptions block scores confidence and forces a test date. Review quarterly; a change needs a Decision Journal row.","CEO · quarterly (rewritten annually)"),
 ("Market Map","Bottom-up TAM / SAM / SOM for Indian fuel distribution, segment scoring, and penetration scenarios priced off the per-GSTIN metric.","TAM = universe × GSTINs per firm × annual price. SOM rows drive off the penetration % you type. Every source number needs a citation and a VERIFY LIVE date.","CEO · half-yearly"),
 ("Unit Economics","CAC, gross margin, payback, LTV and contribution per paying GSTIN, built from real VSYST cost lines.","Everything below the input block is computed. Change only the green cells. CAC is founder time × day rate ÷ close rate.","CEO with CA · quarterly"),
 ("Runway & Scenarios","24-month cash projection in three cases (Base / Bear / Bull) plus the trigger ladder that pre-commits what happens at 12/9/6/3 months of runway.","Pick the scenario in B6; the projection reads its drivers. Runway = closing cash ÷ trailing-3-month net burn. The COO workbook owns the weekly 13-week view; this one owns the scenario.","CEO · monthly, and on any cash surprise"),
 ("Cap Table & Dilution","Current cap table plus a two-round dilution model (seed, Series A) with the pre- vs post-money ESOP pool switch.","Type the round inputs in the green cells; ownership columns recompute. The pool-timing toggle in B30 is the single most expensive cell in this workbook.","CEO with CS · on any change"),
 ("ESOP Pool","Pool size, grant register, vesting maths and the grant guideline by level.","Vested = options × elapsed months ÷ total months, zero before the cliff. Strike and 11UA valuation date come from the CS.","CEO with CS · on every grant"),
 ("CEO Dashboard","The CEO's monthly instrument panel — deliberately different from the COO's weekly KPI Scorecard: fewer rows, longer horizon, strategy and cash first.","Fill one column per month. Status is computed from Direction, Goal and Red line. Read it before the board pack, not after.","CEO · monthly by the 5th"),
 ("Annual Plan & OKRs","The year's theme and bet, 3–5 objectives with key results, owners, baselines, targets and confidence.","% done = (current − baseline) ÷ (target − baseline). Confidence is typed, not computed — that is the point.","CEO · set annually, scored quarterly"),
 ("Decision Journal","Every decision worth remembering: type, reversibility, options, reasoning, what would change your mind, and the outcome scored later.","Verdict compares decision quality against outcome so you stop rewarding luck. Review the open rows every quarter.","CEO · at the decision, reviewed quarterly"),
 ("Board Calendar","Board meeting dates with the statutory gap computed, plus the notice / pack / quorum / minutes trail and the director filings the CEO personally signs.","Gap (days) is computed between consecutive meetings — watch the §173 limit. Dates and forms are VERIFY LIVE with the CS.","CEO with CS · set in April, checked monthly"),
 ("Pipeline","The CEO's own founder-led sales pipeline — named dealers, districts, stage, next step, weighted value and the objection that is actually blocking.","Weighted ₹ = potential × probability. The Next step / Next date pair is the only discipline that matters; a blank one is a dead deal.","CEO · updated after every call"),
 ("Comp Bands","Salary and equity bands by role and level, plus an offer calculator that keeps offers inside the band.","Mid is the anchor; min/max are ±. The offer block flags anything outside the band. Bands are VERIFY LIVE against market each year.","CEO · annually, and before any offer"),
 ("Risk & Pre-Mortem","The pre-mortem — it is 24 months from now and VSYST failed, why? — scored by likelihood × impact, with the early-warning signal and the owner.","Score = Likelihood × Impact. Sort descending; the top three get a mitigation with a date. The COO workbook's Risk Register owns operational risk; this owns existential risk.","CEO · quarterly, and after any near-miss"),
]
r = 5
for i, (tab, what, how, own) in enumerate(readme, start=1):
    put(ws, r, [i, tab, what, how, own]); r += 1
r += 1
ws.cell(row=r, column=1, value="Colour legend").font = Font(sz=10, b=True, color=INK); r += 1
for fill, label in [(EX_FILL, "Example data — grey italic. Overwrite it; it is there to show the shape, and the numbers are illustrative only."),
                    (FM_FILL, "Computed cell — a live formula. Do not type over it."),
                    (IN_FILL, "Input cell — this is where you type."),
                    (SEC_FILL, "Section band — a heading inside the sheet.")]:
    ws.cell(row=r, column=1).fill = fill; ws.cell(row=r, column=1).border = BOX
    ws.cell(row=r, column=2, value=label).font = Font(sz=9, color=MUTED)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True); r += 1
r += 1
ws.cell(row=r, column=1, value="Honesty rule: every rupee figure, statutory date, threshold and market-size input in this workbook is illustrative or was true on 2026-08-25. Anything a real decision turns on is VERIFY LIVE — confirm with the CA, the CS or the source before acting on it.").font = Font(sz=9, i=True, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True); ws.row_dimensions[r].height = 28

# ─────────────────────────────────────────────── Strategy Kernel
ws = sheet("Strategy Kernel", "Strategy Kernel — diagnosis, guiding policy, coherent action (Rumelt) + the Playing-to-Win cascade",
  "Fills C04 and C05. Rewritten annually, reviewed quarterly. A change to the diagnosis is a one-way door — log it in the Decision Journal. Example text is illustrative; argue with it, do not inherit it.",
  [30, 62, 40, 18, 16, 16, 22])
band(ws, 4, "A. The kernel", 7)
header(ws, 5, ["Element", "Our answer (write it so it could be wrong)", "Why we believe it — evidence", "Confidence", "Set on", "Review on", "Owner"])
ws.freeze_panes = "B6"
kern = [
 ("Diagnosis — what is actually going on",
  "Indian fuel dealers do not lack software; they lack a reason to change a paper process that already works for them. The buyer is a busy owner with low software trust, a real switching cost in his ledger, and an accountant who prefers Tally. The binding constraint is adoption friction and district-level trust, not features.",
  "Field conversations; the affordability and education-gap research in docs/learning/company.", 0.6),
 ("Guiding policy — the approach that beats the constraint",
  "Win one district at a time by removing the dealer's single most painful reconciliation, and let dealer-to-dealer proof carry us to the next dealer. Sell trust, not features. Never make the dealer's ledger our experiment.",
  "Follows from the diagnosis; consistent with the per-GSTIN, web-only pricing decision.", 0.6),
 ("Coherent action 1", "Founder-led sales in one district until 10 paying GSTINs, with a written script and objection bank (C10).", "", 0.7),
 ("Coherent action 2", "Onboarding measured as time-to-first-invoice, not as a signup count.", "", 0.7),
 ("Coherent action 3", "Ledger correctness treated as a company-level incident, above any feature.", "", 0.8),
 ("Coherent action 4", "Annual prepay pricing per GSTIN on the web, funding the next district's travel.", "", 0.5),
 ("What we are deliberately NOT doing", "No pan-India launch, no free tier, no bespoke builds for one dealer, no in-app purchase, no OMC-scale pursuit that costs more than one founder-quarter.", "", 0.7),
]
r = 6
for name, ans, ev, conf in kern:
    put(ws, r, [name, ans, ev, conf, date(2026,8,25), date(2026,11,25), "CEO"], example=True,
        fmt={4: PCT, 5: "dd mmm yyyy", 6: "dd mmm yyyy"})
    ws.row_dimensions[r].height = 46
    r += 1
r += 1
band(ws, r, "B. Playing to Win — the choice cascade (Lafley & Martin)", 7); r += 1
header(ws, r, ["Question", "Our choice", "What must be true for this to work", "Confidence", "Set on", "Review on", "Owner"]); r += 1
casc = [
 ("What is our winning aspiration?", "The default operating system for Indian fuel distribution, starting with Chhattisgarh."),
 ("Where will we play?", "Petrol-pump dealers and bulk-diesel operators with ≥1 B2B credit customer, in Chhattisgarh districts first. Not lubricant-only distributors yet. Not pan-India."),
 ("How will we win?", "Mobile-first for the dealer, GST-native, and the only system that holds the dealer↔customer rate confirmation and ledger in one auditable trail."),
 ("What capabilities must be in place?", "Founder-led district selling; onboarding that reaches first invoice in days; ledger correctness; low-bandwidth Android quality; Hindi-first UX."),
 ("What management systems are required?", "The COO course's operating system — cadence, scorecard, SOPs, compliance calendar — plus this workbook's CEO Dashboard and Decision Journal."),
]
for q, a in casc:
    put(ws, r, [q, a, "", 0.6, date(2026,8,25), date(2026,11,25), "CEO"], example=True,
        fmt={4: PCT, 5: "dd mmm yyyy", 6: "dd mmm yyyy"})
    ws.row_dimensions[r].height = 42; r += 1
r += 1
band(ws, r, "C. Assumption register — what would prove the strategy wrong", 7); r += 1
header(ws, r, ["Assumption we are betting on", "If it is false, what breaks", "The smallest test", "Confidence", "Test by", "Tested on", "Verdict / what we learned"]); r += 1
assump = [
 ("A dealer will pay per GSTIN for this, annually, on a web page", "The whole revenue model; pricing metric and billing channel both", "Ask 10 qualified dealers for a signed annual order at list price", 0.5),
 ("The dealer's B2B customers will use the customer app", "Retention and the two-sided ledger asset; the moat argument", "Measure customer-side confirmations per active dealer for 60 days", 0.45),
 ("One district's dealers refer other dealers", "The whole go-to-market; CAC stays high without it", "Track referral source on every deal for two quarters", 0.4),
 ("Reconciliation pain is the top-3 pain for the owner, not the accountant", "The positioning and the demo order", "10 switch interviews (C09), coded for stated pain rank", 0.55),
 ("We can onboard a dealer to first invoice in under 7 days without a founder on site", "Scalability past one district; the first sales hire's viability", "Onboard 3 dealers remotely and time it", 0.35),
]
first_assump = r
for a, b_, c_, conf in assump:
    put(ws, r, [a, b_, c_, conf, date(2026,11,30), "", ""], example=True, fmt={4: PCT, 5: "dd mmm yyyy", 6: "dd mmm yyyy"})
    ws.row_dimensions[r].height = 34; r += 1
ws.conditional_formatting.add(f"D{first_assump}:D{r-1}",
    CellIsRule(operator="lessThan", formula=["0.5"], fill=PatternFill("solid", fgColor="FEE2E2")))
r += 1
band(ws, r, "D. Strategy review log — a change here is a one-way door", 7); r += 1
header(ws, r, ["Date", "What changed", "The evidence that forced it", "Persist / pivot", "Decision Journal ID", "Board informed on", "Owner"]); r += 1
put(ws, r, [date(2026,8,25), "Kernel written for the first time", "Course lesson 05 exercise", "n/a", "DJ-001", "", "CEO"], example=True,
    fmt={1: "dd mmm yyyy", 6: "dd mmm yyyy"})

# ─────────────────────────────────────────────── Market Map
ws = sheet("Market Map", "Market Map — bottom-up TAM / SAM / SOM, segment scoring and penetration scenarios",
  "Fills C06. Every universe count below is an INPUT you must source and date — replace the illustrative figures and cite PPAC / MoPNG / OMC disclosures. VERIFY LIVE before any of this reaches a deck or a board pack.",
  [34, 16, 16, 16, 16, 16, 34])
band(ws, 4, "A. Pricing inputs (from the per-GSTIN decision — see C11)", 7)
header(ws, 5, ["Input", "Value", "Unit", "", "", "", "Source / note"])
ws.freeze_panes = "A6"
put(ws, 6, ["List price per GSTIN per month", 1799, "₹", None, None, None, "Matches C11 and lessons 05/08. VERIFY LIVE."], example=True, fmt={2: RS}, input_cols=(2,))
put(ws, 7, ["Annual price per GSTIN", "=B6*12", "₹", None, None, None, "Annual prepay is the default plan"], formula_cols=(2,), fmt={2: RS})
put(ws, 8, ["Average GSTINs per dealer firm", 1.3, "count", None, None, None, "Multi-GSTIN dealers are the upside. Illustrative."], example=True, input_cols=(2,))
put(ws, 9, ["Average annual revenue per firm (ARPA)", "=B7*B8", "₹", None, None, None, "Drives every number below"], formula_cols=(2,), fmt={2: RS})
band(ws, 11, "B. The universe — replace every figure and cite it", 7)
header(ws, 12, ["Universe", "Count", "Source", "As of", "Addressable %", "Addressable count", "Why this filter"])
uni = [
 ("Retail fuel outlets in India (all OMCs)", 90000, "PPAC / OMC disclosures — VERIFY LIVE", "2026", 0.15, "Dealers with ≥1 B2B credit customer and enough volume to feel reconciliation pain"),
 ("Retail fuel outlets in Chhattisgarh", 1400, "State OMC data — VERIFY LIVE", "2026", 0.25, "Home state; founder can reach them physically"),
 ("Bulk diesel / doorstep delivery operators (India)", 6000, "MoPNG / industry press — VERIFY LIVE", "2026", 0.3, "B2B by definition; highest reconciliation pain"),
 ("Lubricant distributors (India)", 12000, "Industry estimate — VERIFY LIVE", "2026", 0.1, "Different workflow; a later segment, not now"),
]
r = 13
for n, c, s, a, pct, why in uni:
    put(ws, r, [n, c, s, a, pct, f"=B{r}*E{r}", why], example=True, fmt={2: "#,##0", 5: PCT, 6: "#,##0"}, formula_cols=(6,))
    r += 1
r += 1
band(ws, r, "C. TAM / SAM / SOM", 7); r += 1
header(ws, r, ["Layer", "Firms", "Basis", "", "", "Annual ₹", "What it means"]); r += 1
tam_row = r
put(ws, r, ["TAM — every fuel retail outlet in India", "=B13", "Total universe, all outlets", None, None, f"=B{r}*$B$9", "The number for a deck, not for a plan"], formula_cols=(2,6), fmt={2:"#,##0", 6: RS}); r += 1
put(ws, r, ["SAM — India outlets that fit the ICP", "=F13+F15", "Addressable outlets + bulk operators", None, None, f"=B{r}*$B$9", "Who we could sell to with today's product"], formula_cols=(2,6), fmt={2:"#,##0", 6: RS}); r += 1
put(ws, r, ["SOM — Chhattisgarh, addressable", "=F14", "Home-state addressable outlets", None, None, f"=B{r}*$B$9", "The only number the next 18 months are about"], formula_cols=(2,6), fmt={2:"#,##0", 6: RS}); r += 1
som_row = r - 1
r += 1
band(ws, r, "D. Penetration scenarios against SOM", 7); r += 1
header(ws, r, ["Scenario", "Penetration of SOM", "Paying firms", "Paying GSTINs", "Annual ₹", "Monthly ₹", "What it takes"]); r += 1
scen = [("Toehold", 0.01, "One district, founder-led"), ("Beachhead", 0.03, "Three districts, one sales hire"), ("Home-state leader", 0.10, "Statewide, a real sales team and support function"), ("Default alive", 0.05, "Whatever penetration covers total costs — solve for it")]
for name, p, takes in scen:
    put(ws, r, [name, p, f"=ROUND($B${som_row}*B{r},0)", f"=ROUND(C{r}*$B$8,0)", f"=C{r}*$B$9", f"=E{r}/12", takes],
        formula_cols=(3,4,5,6), fmt={2: PCT, 3: "#,##0", 4: "#,##0", 5: RS, 6: RS}, input_cols=(2,))
    r += 1
r += 1
band(ws, r, "E. Segments — score before you choose", 7); r += 1
header(ws, r, ["Segment", "Pain intensity 1-5", "Ability to pay 1-5", "Reachability 1-5", "Product fit today 1-5", "Score (product)", "Verdict"]); r += 1
segs = [("Petrol-pump dealer with B2B credit book", 5, 4, 4, 5, "Primary — this is where to play"),
        ("Bulk diesel / doorstep delivery operator", 5, 4, 3, 4, "Secondary — same pain, harder to reach"),
        ("Petrol-pump dealer, retail-only, no credit customers", 2, 3, 4, 2, "Disqualify — no reconciliation pain to sell against"),
        ("Lubricant distributor", 3, 3, 2, 2, "Later — different workflow, needs product work"),
        ("Fleet / transporter (the dealer's customer)", 4, 3, 2, 3, "Not a buyer — rides free on the dealer's tenant by design")]
seg_first = r
for n, a, b_, c_, d_, v in segs:
    put(ws, r, [n, a, b_, c_, d_, f"=B{r}*C{r}*D{r}*E{r}", v], example=True, formula_cols=(6,))
    r += 1
ws.conditional_formatting.add(f"F{seg_first}:F{r-1}", ColorScaleRule(
    start_type="min", start_color="FEE2E2", end_type="max", end_color="DCFCE7"))
r += 1
band(ws, r, "F. The real alternatives (what we are actually replacing)", 7); r += 1
header(ws, r, ["Alternative", "What it costs the dealer today", "Why he keeps it", "Where it fails him", "Our wedge", "Threat 1-5", "Note"]); r += 1
alts = [("Paper Daily Sales Register", "Hours per day; variance found at month end", "It has never failed him; his staff know it", "No credit visibility, no dispute trail", "The rate-confirmation trail", 5),
        ("Tally + the accountant", "Annual licence + accountant fees", "GST filings depend on it", "Nothing operational; days behind reality", "We feed Tally, we don't fight it", 4),
        ("WhatsApp for orders and rates", "Free", "Every customer already has it", "No record, endless rate disputes", "Confirmed rates with a timestamp", 5),
        ("Excel credit tracker", "Free", "He built it himself", "Breaks past ~30 customers; one file, one person", "Multi-user ledger with lineage", 3),
        ("Another billing app / ATG vendor software", "₹ varies", "Bundled with hardware", "Not built for the B2B credit workflow", "Two-sided, dealer↔customer", 3)]
for a, cost, keep, fails, wedge, t in alts:
    put(ws, r, [a, cost, keep, fails, wedge, t, ""], example=True); r += 1

# ─────────────────────────────────────────────── Unit Economics
ws = sheet("Unit Economics", "Unit Economics — CAC, gross margin, payback, LTV and contribution per paying GSTIN",
  "Fills C12. Green cells are inputs; blue cells compute. At pre-revenue every figure here is a hypothesis — write the date you last tested each one. LTV/CAC is the least trustworthy number on this sheet; payback is the most.",
  [40, 18, 14, 46, 16, 16, 16])
band(ws, 4, "A. Revenue inputs", 7)
header(ws, 5, ["Input", "Value", "Unit", "How to get it honestly", "", "", ""])
ws.freeze_panes = "A6"
rows = [("Price per GSTIN per month", 1799, "₹", "From C11. Use list price, not the discounted one."),
        ("GSTINs per paying firm", 1.3, "count", "Actual average across signed dealers."),
        ("Monthly revenue per firm (ARPA)", "=B6*B7", "₹", "Computed."),
        ("Annual revenue per firm", "=B8*12", "₹", "Computed.")]
r = 6
for n, v, u, h in rows:
    isf = isinstance(v, str) and v.startswith("=")
    put(ws, r, [n, v, u, h], fmt={2: RS}, formula_cols=(2,) if isf else (), input_cols=() if isf else (2,))
    if not isf: ws.cell(row=r, column=2).font = Font(i=True, color="808080")
    r += 1
r += 1
band(ws, r, "B. Cost to serve one dealer tenant, per month", 7); r += 1
header(ws, r, ["Cost line", "₹ / tenant / month", "Basis", "Note", "", "", ""]); r += 1
cts_first = r
costs = [("AWS compute + storage + bandwidth", 120, "Total AWS ÷ active tenants", "Falls with scale; re-measure every quarter"),
         ("MongoDB Atlas", 90, "Cluster cost ÷ active tenants", "Watch the per-tenant document growth"),
         ("SMS / OTP (2Factor) — driver + rate confirmations", 210, "Messages × rate", "Scales with the dealer's transaction volume, not with headcount"),
         ("Push notifications / Firebase", 15, "Flat", ""),
         ("Payment gateway fees on subscription collection", 40, "≈1.7% of ARPA", "Easebuzz rate; price-derived — recompute if C11 price changes. VERIFY LIVE"),
         ("Support cost (human minutes)", 400, "Support hours × loaded cost ÷ tenants", "The one that decides whether this business scales"),
         ("Onboarding amortised over 24 months", 250, "Onboarding cost ÷ 24", "Move to CAC instead if you prefer; be consistent")]
for n, v, b_, note in costs:
    put(ws, r, [n, v, b_, note], example=True, fmt={2: RS}, input_cols=(2,)); r += 1
cts_last = r - 1
put(ws, r, ["Total cost to serve, per tenant per month", f"=SUM(B{cts_first}:B{cts_last})", "", "Computed"], formula_cols=(2,), fmt={2: RS}); cts_total = r; r += 1
put(ws, r, ["Gross margin per tenant per month", f"=B8-B{cts_total}", "", "Computed"], formula_cols=(2,), fmt={2: RS}); gm_abs = r; r += 1
put(ws, r, ["Gross margin %", f"=IF(B8=0,\"\",B{gm_abs}/B8)", "", "SaaS benchmark is 70–80%+. Below 60% means support or SMS is eating the business."], formula_cols=(2,), fmt={2: PCT}); gm_pct = r; r += 2
band(ws, r, "C. Cost to acquire one dealer (founder-led, district by district)", 7); r += 1
header(ws, r, ["Input", "Value", "Unit", "How to get it honestly", "", "", ""]); r += 1
cac_first = r
cac = [("Founder days spent per closed deal", 6, "days", "Count travel, demos, follow-ups and the paperwork. Be honest."),
       ("Loaded cost of a founder day", 4000, "₹", "Salary ÷ working days. Use the real number even if you don't pay yourself yet — it is a real cost."),
       ("Travel + hospitality per deal", 3500, "₹", "Fuel, stay, food, the dealer's tea."),
       ("Marketing / collateral per deal", 500, "₹", "Printing, demo device, association fees amortised."),
       ("Deals closed ÷ qualified conversations", 0.2, "ratio", "The honest close rate, counted over 20+ conversations.")]
for n, v, u, h in cac:
    put(ws, r, [n, v, u, h], example=True, fmt={2: RS if u == "₹" else ("0.00" if u=="ratio" else "0.0")}, input_cols=(2,)); r += 1
put(ws, r, ["CAC per closed dealer", f"=(B{cac_first}*B{cac_first+1}+B{cac_first+2}+B{cac_first+3})/B{cac_first+4}", "₹",
            "Founder time is the dominant term. If this frightens you, that is the correct reaction — it is why lesson 07 insists on a script."], formula_cols=(2,), fmt={2: RS})
cac_row = r; r += 2
band(ws, r, "D. The four numbers that decide the business", 7); r += 1
header(ws, r, ["Metric", "Value", "Target", "Reading", "", "", ""]); r += 1
put(ws, r, ["Gross margin %", f"=B{gm_pct}", 0.75, "Below target → attack SMS cost and support minutes before attacking price."], formula_cols=(2,), fmt={2: PCT, 3: PCT}); r += 1
put(ws, r, ["CAC payback (months)", f"=IF(B{gm_abs}<=0,\"never\",B{cac_row}/B{gm_abs})", 12, "Months of gross margin to repay acquisition. Under 12 is healthy for SMB SaaS; over 24 means the motion is wrong, not the price."], formula_cols=(2,), fmt={2: "0.0", 3: "0"}); pay_row = r; r += 1
put(ws, r, ["Monthly logo churn %", 0.02, 0.015, "Typed, from actuals. Anything above 3%/month makes LTV meaningless."], example=True, fmt={2: PCT, 3: PCT}, input_cols=(2,)); churn_row = r; r += 1
put(ws, r, ["Expected lifetime (months)", f"=IF(B{churn_row}=0,\"\",1/B{churn_row})", "", "1 ÷ monthly churn."], formula_cols=(2,), fmt={2: "0"}); life_row = r; r += 1
put(ws, r, ["LTV (gross margin basis)", f"=B{gm_abs}*B{life_row}", "", "Gross margin, never revenue. Revenue-based LTV is a lie you tell yourself."], formula_cols=(2,), fmt={2: RS}); ltv_row = r; r += 1
put(ws, r, ["LTV ÷ CAC", f"=IF(B{cac_row}=0,\"\",B{ltv_row}/B{cac_row})", 3, "Rule of thumb only. With fewer than ~30 customers this ratio is noise — trust payback instead."], formula_cols=(2,), fmt={2: "0.0", 3: "0.0"}); r += 1
put(ws, r, ["Paying firms needed to cover total monthly costs", f"=IF(B{gm_abs}<=0,\"never\",ROUNDUP(150000/B{gm_abs},0))", "", "Replace 150000 with this month's total operating cost from Runway & Scenarios. This is the 'default alive' number."], formula_cols=(2,), fmt={2: "#,##0"})
ws.conditional_formatting.add(f"B{pay_row}", CellIsRule(operator="greaterThan", formula=["24"], fill=PatternFill("solid", fgColor="FEE2E2")))

# ─────────────────────────────────────────────── Runway & Scenarios
ws = sheet("Runway & Scenarios", "Runway & Scenarios — 24-month cash in three cases, and the trigger ladder",
  "Fills C13. The COO workbook owns the weekly 13-week cash view; this sheet owns the scenario and the pre-committed trigger. Change the scenario in B5 and the whole projection re-reads its drivers. Figures illustrative — VERIFY LIVE against the bank.",
  [26, 14, 14, 14, 14, 14, 14, 14, 40])
band(ws, 4, "A. Scenario switch and drivers", 9)
ws["A5"] = "Active scenario →"; ws["A5"].font = Font(b=True)
ws["B5"] = "Base"; ws["B5"].fill = IN_FILL; ws["B5"].border = BOX; ws["B5"].font = Font(b=True)
ws["C5"] = "Type Base, Bear or Bull. Everything below reads this cell."; ws["C5"].font = Font(sz=9, i=True, color=MUTED)
header(ws, 7, ["Driver", "Bear", "Base", "Bull", "Active", "", "", "", "How to set it honestly"])
ws.freeze_panes = "A8"
drv = [("Opening cash today (₹)", 2500000, 2500000, 2500000, RS, "Bank + FD + anything liquid, today. One number, no optimism."),
       ("New paying firms per month", 0.5, 1.5, 3, "0.0", "Your actual close rate × conversations you can physically hold."),
       ("Annual revenue per firm (₹)", 21588, 28065, 32382, RS, "From Unit Economics B9 (1799 x 12 x GSTINs/firm)."),
       ("Monthly logo churn", 0.04, 0.02, 0.01, PCT, "From actuals once you have any."),
       ("Fixed monthly costs (₹)", 190000, 165000, 165000, RS, "Salaries, cloud, tools, CA/CS, rent. The bear case is higher because problems cost money."),
       ("Variable cost per firm per month (₹)", 1400, 1125, 900, RS, "Base = C12 §B cost to serve (₹1,125). Bear/bull are scenario assumptions."),
       ("Hiring: month of first hire", 99, 12, 8, "0", "99 = not in this scenario."),
       ("Hiring: added monthly cost (₹)", 0, 45000, 60000, RS, "Fully loaded, including PF/ESI where applicable.")]
r = 8
first_drv = r
for n, be, ba, bu, fmt, how in drv:
    put(ws, r, [n, be, ba, bu, f'=IF($B$5="Bear",B{r},IF($B$5="Bull",D{r},C{r}))', None, None, None, how],
        example=True, fmt={2: fmt, 3: fmt, 4: fmt, 5: fmt}, formula_cols=(5,), input_cols=(2,3,4))
    r += 1
OC, NEW, ARPA, CHURN, FIX, VAR, HM, HC = [first_drv + i for i in range(8)]
r += 1
band(ws, r, "B. The floor — cash that is not spendable (C13 rule 11)", 9); r += 1
header(ws, r, ["Reserve", "₹", "", "", "", "", "", "", "How to set it honestly"]); r += 1
OPRES = r
put(ws, r, ["Operating reserve", 300000, None, None, None, None, None, None,
    "One bad month is not an emergency. Roughly two months of fixed cost."],
    example=True, fmt={2: RS}, input_cols=(2,)); r += 1
EXRES = r
put(ws, r, ["Exit reserve — ring-fenced, not spendable", 450000, None, None, None, None, None, None,
    "What it would cost to end the company properly, today: closing accounts, strike-off (STK-2 via C-PACE) or IBC s.59, unexpired prepayments refunded, vendor and settlement dues, full-and-final for anyone employed, per-tenant data export and verified deletion under DPDP s.8, plus a 25% buffer. Drivers in lesson 18 §10.6. VERIFY LIVE with the CA and the CS — re-price at the annual plan, after every hire, and on signing any annual prepayment."],
    example=True, fmt={2: RS}, input_cols=(2,)); r += 1
FLOOR = r
put(ws, r, ["Total floor", f"=B{OPRES}+B{EXRES}", None, None, None, None, None, None,
    "Column I below reads runway net of the exit reserve. The 3-month rung may not spend it."],
    formula_cols=(2,), fmt={2: RS}); r += 1
BREACH = r
put(ws, r, ["Month closing cash first falls below the floor", None, None, None, None, None, None, None,
    "If this lands inside 24 months, the ladder is not theoretical. Read it with section D."],
    formula_cols=(2,)); r += 1
r += 1
band(ws, r, "C. 24-month projection (reads the active column)", 9); r += 1
header(ws, r, ["Month #", "Paying firms", "Revenue ₹", "Fixed cost ₹", "Variable cost ₹", "Total cost ₹", "Net ₹", "Closing cash ₹", "Runway (months, net of exit reserve)"]); r += 1
proj_first = r
for m in range(1, 25):
    prev = r - 1
    firms = f"=$E${NEW}" if m == 1 else f"=MAX(0,B{prev}*(1-$E${CHURN})+$E${NEW})"
    rev = f"=B{r}*$E${ARPA}/12"
    fixc = f"=$E${FIX}+IF({m}>=$E${HM},$E${HC},0)"
    varc = f"=B{r}*$E${VAR}"
    tot = f"=D{r}+E{r}"
    net = f"=C{r}-F{r}"
    close = f"=$E${OC}+G{r}" if m == 1 else f"=H{prev}+G{r}"
    runw = f'=IF(AVERAGE(G{max(proj_first, r-2)}:G{r})>=0,"n/a — cash positive",IF(H{r}-$B${EXRES}<=0,0,(H{r}-$B${EXRES})/-AVERAGE(G{max(proj_first, r-2)}:G{r})))'
    put(ws, r, [m, firms, rev, fixc, varc, tot, net, close, runw],
        formula_cols=(2,3,4,5,6,7,8,9), fmt={2: "0.0", 3: RS, 4: RS, 5: RS, 6: RS, 7: RS, 8: RS, 9: "0.0"})
    r += 1
proj_last = r - 1
ws.cell(row=BREACH, column=2).value = (
    f'=IF(COUNTIF(H{proj_first}:H{proj_last},"<"&B{FLOOR})=0,"never within 24 months",'
    f'SUMPRODUCT(MIN((H{proj_first}:H{proj_last}<B{FLOOR})*A{proj_first}:A{proj_last}'
    f'+(H{proj_first}:H{proj_last}>=B{FLOOR})*9999)))')
ws.cell(row=BREACH, column=2).number_format = "0"
ws.conditional_formatting.add(f"H{proj_first}:H{proj_last}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE2E2")))
ws.conditional_formatting.add(f"I{proj_first}:I{proj_last}", CellIsRule(operator="lessThan", formula=["6"], fill=PatternFill("solid", fgColor="FEF3C7")))
r += 1
band(ws, r, "D. The trigger ladder — decided now, in the calm, not later in the panic", 9); r += 1
header(ws, r, ["Runway remaining", "Status", "What the CEO is obliged to do — no discretion", "", "", "", "", "", "Who must be told"]); r += 1
lad = [("12 months", "Green", "Normal operations. Review this ladder monthly and confirm the cash figure against the bank statement personally.", "Board / advisors at the quarterly"),
       ("9 months", "Amber", "Write the plan you would execute at 6 months, and name the date you would start it. Freeze any new recurring cost above ₹10,000/month. Start the funding conversation you have been postponing (grants, prepay push, or investors).", "Co-founders, same week"),
       ("6 months", "Red", "Execute the 9-month plan. Cut discretionary spend to zero. Convert every willing customer to annual prepay at a discount. Stop all work not on the shortest path to revenue. If raising, the process must already be running.", "Co-founders same day; board within a week"),
       ("3 months", "Critical", "Founder salaries to statutory minimum or zero. The exit reserve is NOT available to this rung — spending it converts a company that can still stop well into one that can only stop badly. Talk to the CA about statutory dues sequencing so nothing becomes personal liability. Have the honest conversation with the team before they hear it elsewhere. Decide, in writing, the date on which you would stop.", "Everyone, immediately"),
       ("Cash-flow positive", "Earned", "Rebuild to 12 months of runway before adding any recurring cost. Then, and only then, spend on growth.", "Board at the next meeting")]
for a, s, act, who in lad:
    put(ws, r, [a, s, act, None, None, None, None, None, who], example=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 44
    r += 1

# ─────────────────────────────────────────────── Cap Table & Dilution
ws = sheet("Cap Table & Dilution", "Cap Table & Dilution — the current table and a two-round model",
  "Fills C16. Illustrative shares and valuations. Every real change goes through the CS and the statutory registers — this sheet is a model, never the record. The pool-timing cell is the most expensive single input on it.",
  [30, 16, 16, 16, 16, 16, 16, 16, 34])
band(ws, 4, "A. Cap table today", 9)
header(ws, 5, ["Holder", "Class", "Shares", "% fully diluted", "", "", "", "", "Note"])
ws.freeze_panes = "A6"
cap = [("Founder-Director 1 (CEO)", "Equity", 400000),
       ("Founder-Director 2 (domain)", "Equity", 300000),
       ("Founder-Director 3", "Equity", 300000),
       ("ESOP pool (unallocated)", "Reserved", 0)]
r = 6; cap_first = r
for h, cl, sh in cap:
    put(ws, r, [h, cl, sh, f"=IF($C${cap_first+len(cap)}=0,\"\",C{r}/$C${cap_first+len(cap)})", None,None,None,None,
                "Founder shares should be subject to a vesting / buy-back arrangement — see C02 and the CS."],
        example=True, fmt={3: "#,##0", 4: PCT}, formula_cols=(4,), input_cols=(3,))
    r += 1
put(ws, r, ["Total", "", f"=SUM(C{cap_first}:C{r-1})", f"=SUM(D{cap_first}:D{r-1})"], formula_cols=(3,4), fmt={3: "#,##0", 4: PCT})
cap_total = r; r += 2
band(ws, r, "B. Round model — inputs", 9); r += 1
header(ws, r, ["Input", "Seed", "Series A", "", "", "", "", "", "Note"]); r += 1
ri = r
rinp = [("Pre-money valuation (₹)", 60000000, 300000000, RS, "The number everyone argues about and that matters less than the terms below."),
        ("Amount raised (₹)", 15000000, 60000000, RS, ""),
        ("Target ESOP pool after the round", 0.10, 0.12, PCT, "The pool is topped up to this % of post-money."),
        ("Pool created PRE-money? (1 = yes)", 1, 1, "0", "1 = founders alone bear the pool dilution (the market standard, and the expensive one). 0 = everyone shares it. Read lesson 08 §12 before typing here."),
        ("Liquidation preference (x)", 1, 1, "0.0", "1x non-participating is the founder-friendly norm in India. VERIFY in the term sheet."),
        ("Participating? (1 = yes)", 0, 0, "0", "Participating preference double-dips on exit. Push back hard.")]
for n, s, a, fmt, note in rinp:
    put(ws, r, [n, s, a, None,None,None,None,None, note], example=True, fmt={2: fmt, 3: fmt}, input_cols=(2,3)); r += 1
PRE, AMT, POOL, PREPOOL, LP, PART = [ri + i for i in range(6)]
r += 1
band(ws, r, "C. Round model — computed outcome", 9); r += 1
header(ws, r, ["Output", "Seed", "Series A", "", "", "", "", "", "Reading"]); r += 1
o = r
outs = [("Post-money valuation (₹)", f"=B{PRE}+B{AMT}", f"=C{PRE}+C{AMT}", RS, "Post = pre + amount. Nothing subtle here."),
        ("Investor ownership after round", f"=B{AMT}/B{o}", f"=C{AMT}/C{o}", PCT, "What the money buys."),
        ("ESOP pool after round", f"=B{POOL}", f"=C{POOL}", PCT, "Target pool as a share of post-money."),
        ("Everyone else (founders + prior investors)", f"=1-B{o+1}-B{o+2}", f"=1-C{o+1}-C{o+2}", PCT, "The residual. This is the number to watch across rounds."),
        ("Effective price per share vs founders' cost", "", "", "General", "Ask the CS — needs the share count and Rule 11UA valuation."),
        ("Founder group ownership after Seed", f"=B{o+3}", "", PCT, "Assumes no prior outside holders."),
        ("Founder group ownership after Series A", f"=B{o+5}*C{o+3}", "", PCT, "Compounding dilution — this is the number founders consistently under-estimate."),
        ("Investor's minimum exit for a 1x return (₹)", f"=B{AMT}*B{LP}", f"=C{AMT}*C{LP}", RS, "Below this, common shareholders may see nothing."),
        ("Pool dilution borne by founders alone?", f'=IF(B{PREPOOL}=1,"Yes — founders pay for the pool","No — shared with the new investor")', f'=IF(C{PREPOOL}=1,"Yes — founders pay for the pool","No — shared with the new investor")', "General", "On a ₹6 Cr pre-money, a 10% pre-money pool costs the founders roughly ₹60 lakh of value. It is a negotiable term.")]
for n, s, a, fmt, note in outs:
    put(ws, r, [n, s, a, None,None,None,None,None, note], formula_cols=(2,3), fmt={2: fmt, 3: fmt}); r += 1
r += 1
band(ws, r, "D. Terms that matter more than the valuation — fill from the actual term sheet", 9); r += 1
header(ws, r, ["Term", "What was offered", "Founder-friendly benchmark", "Accept / negotiate / refuse", "", "", "", "", "Why it matters"]); r += 1
terms = [("Liquidation preference", "1x non-participating", "Decides who gets paid first, and how much, on any exit."),
         ("Anti-dilution", "Broad-based weighted average", "Full ratchet transfers the whole cost of a down round to founders."),
         ("Board composition", "Founder majority at seed", "This is the term that decides whether you can be fired."),
         ("Reserved matters / protective provisions", "Short list, high thresholds", "A long list means you need consent to run your own company."),
         ("ESOP pool and its timing", "Post-money, or shared", "See row above — real money."),
         ("Founder vesting reset", "Credit for time already served", "Refusing all credit for four years of work is a red flag about the investor."),
         ("Drag-along", "Triggered by a real majority incl. founders", "Can force you to sell."),
         ("Information rights", "Quarterly, standard", "Fine. Monthly audited is not."),
         ("Exclusivity / no-shop", "30–45 days", "Long exclusivity with a slow investor kills your leverage."),
         ("Tag-along / ROFR", "Standard", "Governs secondary sales.")]
for t, bench, why in terms:
    put(ws, r, [t, "", bench, "", None,None,None,None, why], example=True); r += 1

# ─────────────────────────────────────────────── ESOP Pool
ws = sheet("ESOP Pool", "ESOP Pool — pool size, grant register and vesting",
  "Fills C20. Grants for an Indian private company run under the Companies (Share Capital and Debentures) Rules, 2014 and need a board and shareholder resolution — the CS drives it, this sheet only tracks it. Strike price and valuation date: VERIFY LIVE with the CS. Perquisite tax on exercise applies — see lesson 10 §8.",
  [26, 20, 14, 14, 14, 14, 14, 14, 34])
band(ws, 4, "A. Pool", 9)
header(ws, 5, ["Item", "Value", "", "", "", "", "", "", "Note"])
ws.freeze_panes = "A6"
put(ws, 6, ["Pool size (% fully diluted)", 0.07, None,None,None,None,None,None, "7% — matches C20 and lesson 10. 10% is the seed-stage norm; revisit at the round. VERIFY LIVE."], example=True, fmt={2: PCT}, input_cols=(2,))
put(ws, 7, ["Pool size (options)", 75269, None,None,None,None,None,None, "From the cap table."], example=True, fmt={2: "#,##0"}, input_cols=(2,))
put(ws, 8, ["Granted to date", "=SUM(D14:D30)", None,None,None,None,None,None, "Computed from the register below."], formula_cols=(2,), fmt={2: "#,##0"})
put(ws, 9, ["Available to grant", "=B7-B8", None,None,None,None,None,None, "Keep at least 40% of the pool for the next 18 months of hiring."], formula_cols=(2,), fmt={2: "#,##0"})
band(ws, 11, "B. Grant register", 9)
header(ws, 13, ["Grantee", "Role / level", "Grant date", "Options", "Strike ₹", "Vest start", "Cliff (months)", "Total (months)", "Vested today (computed)"])
r = 14
put(ws, r, ["A. Example", "Senior Engineer / L5", date(2026,10,1), 6000, 10, date(2026,10,1), 12, 48,
            f'=IF(TODAY()<EDATE(F{r},G{r}),0,MIN(D{r},ROUND(D{r}*DATEDIF(F{r},TODAY(),"m")/H{r},0)))'],
    example=True, fmt={3: "dd mmm yyyy", 4: "#,##0", 5: RS2, 6: "dd mmm yyyy", 9: "#,##0"}, formula_cols=(9,))
for rr in range(15, 31):
    put(ws, rr, ["", "", "", "", "", "", 12, 48,
                 f'=IF(OR(D{rr}="",F{rr}=""),"",IF(TODAY()<EDATE(F{rr},G{rr}),0,MIN(D{rr},ROUND(D{rr}*DATEDIF(F{rr},TODAY(),"m")/H{rr},0))))'],
        fmt={3: "dd mmm yyyy", 4: "#,##0", 5: RS2, 6: "dd mmm yyyy", 9: "#,##0"}, formula_cols=(9,), input_cols=(1,2,3,4,5,6))
band(ws, 32, "C. Grant guideline by level — set once, deviate only with a written reason", 9)
header(ws, 33, ["Level", "Role examples", "% of fully diluted", "Options (at current pool)", "Cash band ref", "", "", "", "Note"])
r = 34
lv = [("L6 — Function head / first exec", "Head of Sales, CTO", 0.015, "Cap the first exec grant; you will need room for their successor."),
      ("L5 — Senior individual contributor / lead", "Lead engineer, senior CSM", 0.006, ""),
      ("L4 — Mid", "Engineer, implementation lead", 0.003, ""),
      ("L3 — Junior", "Support, junior engineer", 0.001, "Small but real. A grant of zero says something."),
      ("Advisor", "Fuel-industry veteran, SaaS operator", 0.0025, "0.1–0.5% over 2 years with a quarterly vest is the usual shape. VERIFY LIVE.")]
for l, ex, pct, note in lv:
    put(ws, r, [l, ex, pct, f"=ROUND(C{r}*$B$7/$B$6,0)", "see Comp Bands", None,None,None, note], example=True,
        fmt={3: PCT, 4: "#,##0"}, formula_cols=(4,)); r += 1
r += 1
ws.cell(row=r, column=1, value="Liquidity warning: in a bootstrapped company that may never exit, an ESOP is a promise with no market. Say so out loud when you grant it, and consider a written buy-back commitment or a phantom/SAR structure instead — lesson 10 §8.").font = Font(sz=9, i=True, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True); ws.row_dimensions[r].height = 30

# ─────────────────────────────────────────────── CEO Dashboard
ws = sheet("CEO Dashboard", "CEO Dashboard — the monthly instrument panel",
  "Fills C29's companion. Deliberately NOT the COO's weekly KPI Scorecard: fewer rows, a monthly beat, and cash and strategy at the top. Fill one column per month by the 5th. Status computes from Direction, Goal and Red line.",
  [4, 34, 14, 26, 11, 14, 13, 12] + [11]*12 + [12, 12])
ws["H3"] = "Month →"; ws["H3"].font = Font(sz=9, b=True, color=MUTED)
ws["I3"] = date(2026, 9, 1); ws["I3"].number_format = "mmm yyyy"; ws["I3"].fill = IN_FILL
for i in range(1, 12):
    c = ws.cell(row=3, column=9 + i, value=f"=EDATE({get_column_letter(9+i-1)}3,1)")
    c.number_format = "mmm yyyy"; c.fill = FM_FILL
header(ws, 4, ["#", "Metric", "Owner", "Source", "Unit", "Direction", "Goal (Green)", "Red line"] + [""]*12 + ["Last", "Status"])
ws.freeze_panes = "C5"
groups = [
 ("CASH AND SURVIVAL", [
   ("Cash in bank (₹, month end)", "CEO", "Bank statement, personally checked", "₹", "Higher is better", 2000000, 800000),
   ("Net burn (₹/month)", "CEO with CA", "Runway & Scenarios", "₹", "Lower is better", 150000, 250000),
   ("Runway (months, net of exit reserve)", "CEO", "Runway & Scenarios col I", "months", "Higher is better", 12, 6),
   ("Exit reserve still intact (₹)", "CEO", "Runway & Scenarios, ring-fenced line", "₹", "Higher is better", 450000, 450000)]),
 ("REVENUE AND CUSTOMERS", [
   ("Paying GSTINs", "CEO", "Billing system", "count", "Higher is better", 20, 8),
   ("MRR (₹)", "CEO", "Billing system = paying GSTINs × C11 price", "₹", "Higher is better", 36000, 14400),
   ("Net new paying firms this month", "CEO", "Billing system", "count", "Higher is better", 2, 0),
   ("Logo churn this month", "CEO", "Billing system", "%", "Lower is better", 0.01, 0.03),
   ("Districts with ≥1 paying dealer", "CEO", "CRM", "count", "Higher is better", 3, 1)]),
 ("STRATEGY AND MARKET", [
   ("Strategy-bet progress (0–100%)", "CEO", "Annual Plan & OKRs", "%", "Higher is better", 0.75, 0.4),
   ("Assumptions tested this quarter", "CEO", "Strategy Kernel tab C", "count", "Higher is better", 2, 0),
   ("Founder customer conversations this month", "CEO", "Pipeline tab", "count", "Higher is better", 20, 8),
   ("Weighted pipeline (₹ ARR)", "CEO", "Pipeline tab", "₹", "Higher is better", 300000, 100000)]),
 ("PEOPLE", [
   ("Open roles past 60 days", "CEO", "Hiring tracker", "count", "Lower is better", 0, 2),
   ("Regretted attrition (rolling 12m)", "CEO", "HR record", "count", "Lower is better", 0, 1),
   ("Team size (FTE + contractors)", "COO", "Headcount plan", "count", "Higher is better", 5, 3)]),
 ("GOVERNANCE AND TRUST", [
   ("Statutory / board misses this month", "CEO with CS", "Board Calendar", "count", "Lower is better", 0, 1),
   ("Money-affecting product incidents", "CEO", "Incident log (COO T17)", "count", "Lower is better", 0, 1),
   ("Days since the CEO last spoke to a paying dealer", "CEO", "Calendar", "days", "Lower is better", 7, 21)]),
]
r = 5
for gname, metrics in groups:
    band(ws, r, gname, 22); r += 1
    for i, (m, own, src, unit, direc, goal, red) in enumerate(metrics, start=1):
        fmt = RS if unit == "₹" else (PCT if unit == "%" else "#,##0")
        put(ws, r, [None, m, own, src, unit, direc, goal, red], fmt={7: fmt, 8: fmt})
        ws.cell(row=r, column=1, value=r-5)
        for cc in range(9, 21):
            cell = ws.cell(row=r, column=cc); cell.number_format = fmt; cell.border = BOX; cell.fill = IN_FILL
        ws.cell(row=r, column=21, value=f'=IFERROR(LOOKUP(2,1/(I{r}:T{r}<>""),I{r}:T{r}),"")').number_format = fmt
        ws.cell(row=r, column=21).fill = FM_FILL
        ws.cell(row=r, column=22, value=(
            f'=IF(U{r}="","",IF(F{r}="Higher is better",IF(U{r}>=G{r},"GREEN",IF(U{r}<=H{r},"RED","AMBER")),'
            f'IF(U{r}<=G{r},"GREEN",IF(U{r}>=H{r},"RED","AMBER"))))')).fill = FM_FILL
        r += 1
last_row = r - 1
ws.conditional_formatting.add(f"V5:V{last_row}", CellIsRule(operator="equal", formula=['"RED"'], fill=PatternFill("solid", fgColor="FEE2E2")))
ws.conditional_formatting.add(f"V5:V{last_row}", CellIsRule(operator="equal", formula=['"AMBER"'], fill=PatternFill("solid", fgColor="FEF3C7")))
ws.conditional_formatting.add(f"V5:V{last_row}", CellIsRule(operator="equal", formula=['"GREEN"'], fill=PatternFill("solid", fgColor="DCFCE7")))
r += 1
ws.cell(row=r, column=2, value="Rule: a metric that is RED two months running becomes an agenda item at the next board meeting automatically, without a vote. A metric that has not driven a decision in two quarters gets deleted.").font = Font(sz=9, i=True, color=MUTED)

# ─────────────────────────────────────────────── Annual Plan & OKRs
ws = sheet("Annual Plan & OKRs", "Annual Plan & OKRs — the year's bet, the objectives, and the honest confidence",
  "Fills C28. Set annually, scored quarterly. Three to five objectives, never more. Confidence is typed by the owner, not computed — a falling confidence line is the earliest warning the plan has of anything.",
  [6, 40, 20, 14, 14, 14, 12, 12, 12, 34])
band(ws, 4, "A. The year", 10)
put(ws, 5, ["Year", "FY 2026-27"], example=True)
put(ws, 6, ["The one thing that must be true by year end", "Twenty paying GSTINs across three Chhattisgarh districts, renewing, with a written and tested sales script."], example=True)
ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=9)
put(ws, 7, ["The bet we are making", "That district-level proof and dealer-to-dealer referral beats broad marketing for this buyer."], example=True)
ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=9)
put(ws, 8, ["What we are deliberately not doing this year", "No pan-India push, no second product, no OMC pursuit beyond one founder-quarter, no outside capital unless the trigger in C13 fires."], example=True)
ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=9)
for rr in (6,7,8):
    ws.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[rr].height = 30
band(ws, 10, "B. Objectives and key results", 10)
header(ws, 11, ["#", "Objective / Key result", "Owner", "Baseline", "Target", "Current", "% done", "Confidence", "Quarter", "Note / what would make this red"])
ws.freeze_panes = "C12"
okrs = [
 ("O1", "Prove the district motion", None, None, None, None, None, "Q1–Q4", None),
 ("KR", "Paying GSTINs", "CEO", 0, 20, 0, 0.6, "Q4", "Counted as billed and collected, not as signed."),
 ("KR", "Districts with ≥1 paying dealer", "CEO", 0, 3, 0, 0.5, "Q4", ""),
 ("KR", "Written sales script tested on ≥20 conversations", "CEO", 0, 20, 0, 0.7, "Q2", ""),
 ("O2", "Make onboarding survive without a founder", None, None, None, None, None, "Q2–Q4", None),
 ("KR", "Median days from signup to first invoice", "COO", 21, 7, 21, 0.5, "Q3", "Lower is better — invert the % done formula for this row."),
 ("KR", "Dealers onboarded with no founder on site", "COO", 0, 3, 0, 0.4, "Q4", ""),
 ("O3", "Stay default alive", None, None, None, None, None, "Q1–Q4", None),
 ("KR", "Runway at year end (months)", "CEO", 14, 12, 14, 0.6, "Q4", "Below 9 at any point triggers the C13 ladder."),
 ("KR", "Gross margin %", "CEO", 0.55, 0.7, 0.55, 0.45, "Q4", "SMS and support cost are the two levers."),
 ("O4", "Earn the right to be trusted with a dealer's money", None, None, None, None, None, "Q1–Q4", None),
 ("KR", "Money-affecting incidents", "CEO", 1, 0, 0, 0.6, "Q4", "One is too many. This row never gets deleted."),
 ("KR", "Board meetings held on schedule with signed minutes", "CEO", 0, 4, 0, 0.8, "Q4", ""),
]
r = 12
for tag, name, own, base, tgt, cur, conf, q, note in okrs:
    if tag != "KR":
        band(ws, r, f"{tag} — {name}", 10)
        ws.cell(row=r, column=9, value=q).font = Font(sz=9, b=True, color=INK)
        r += 1; continue
    put(ws, r, [tag, name, own, base, tgt, cur,
                f'=IF(OR(D{r}="",E{r}="",E{r}=D{r}),"",(F{r}-D{r})/(E{r}-D{r}))', conf, q, note],
        example=True, fmt={7: PCT, 8: PCT}, formula_cols=(7,), input_cols=(4,5,6,8))
    r += 1
ws.conditional_formatting.add(f"H12:H{r-1}", CellIsRule(operator="lessThan", formula=["0.5"], fill=PatternFill("solid", fgColor="FEE2E2")))
ws.conditional_formatting.add(f"G12:G{r-1}", ColorScaleRule(start_type="num", start_value=0, start_color="FEE2E2", end_type="num", end_value=1, end_color="DCFCE7"))
r += 1
band(ws, r, "C. Quarterly review — what actually happened", 10); r += 1
header(ws, r, ["Q", "What we said we'd do", "What happened", "Score 0–1", "Persist / adjust / kill", "Decision Journal ID", "", "", "", "Lesson learned"]); r += 1
for q in ["Q1", "Q2", "Q3", "Q4"]:
    put(ws, r, [q, "", "", "", "", "", None,None,None, ""], input_cols=(2,3,4,5,6,10), fmt={4: "0.00"}); r += 1

# ─────────────────────────────────────────────── Decision Journal
ws = sheet("Decision Journal", "Decision Journal — the record that turns experience into calibration",
  "Fills C29. Write the entry BEFORE the outcome is known; that is the entire point. Review open rows every quarter. The Verdict column exists to stop you rewarding luck: a good decision with a bad outcome is still a good decision.",
  [9, 12, 38, 13, 12, 34, 34, 34, 11, 12, 30, 12, 22])
header(ws, 4, ["ID", "Date", "The decision", "Door", "Clock", "Options considered (incl. do nothing)", "Chosen because", "What would change my mind", "Confidence", "Review on", "What actually happened", "Outcome", "Verdict (computed)"])
ws.freeze_panes = "C5"
r = 5
put(ws, r, ["DJ-001", date(2026,7,29), "Charge per GSTIN, not per user; bill on the web only, not in-app.",
            "One-way", "Year",
            "(a) per user (b) per transaction (c) % of volume (d) per GSTIN (e) free + services. Do nothing = stay free.",
            "GSTIN is the unit the dealer already thinks in and already pays tax against; it scales with his business without punishing him for adding staff. Web-only billing keeps 15–30% of revenue that would otherwise go to app-store commission.",
            "If >30% of qualified dealers refuse to complete a web payment, or if multi-GSTIN dealers turn out to be <10% of the base.",
            0.75, date(2027,1,31), "", "",
            '=IF(OR(L5="",I5=""),"open",IF(AND(I5>=0.6,L5="Good"),"Good decision, good outcome",IF(AND(I5>=0.6,L5="Bad"),"Good decision, bad outcome — do not change the process",IF(L5="Good","Bad decision, good outcome — do not repeat it","Bad decision, bad outcome — change the process"))))'],
    example=True, fmt={2: "dd mmm yyyy", 9: PCT, 10: "dd mmm yyyy"}, formula_cols=(13,))
ws.row_dimensions[r].height = 70
for rr in range(6, 46):
    put(ws, rr, ["", "", "", "", "", "", "", "", "", "", "", "",
                 f'=IF(C{rr}="","",IF(OR(L{rr}="",I{rr}=""),"open",IF(AND(I{rr}>=0.6,L{rr}="Good"),"Good decision, good outcome",IF(AND(I{rr}>=0.6,L{rr}="Bad"),"Good decision, bad outcome — do not change the process",IF(L{rr}="Good","Bad decision, good outcome — do not repeat it","Bad decision, bad outcome — change the process")))))'],
        fmt={2: "dd mmm yyyy", 9: PCT, 10: "dd mmm yyyy"}, formula_cols=(13,), input_cols=(1,2,3,4,5,6,7,8,9,10,11,12))
ws.cell(row=47, column=3, value="Door: One-way (hard to reverse — slow down, gather more) or Two-way (cheap to reverse — decide fast and move). Clock: Week / Year / Decade. Outcome: type Good or Bad at the review date.").font = Font(sz=9, i=True, color=MUTED)
ws.merge_cells(start_row=47, start_column=3, end_row=47, end_column=8); ws.cell(row=47, column=3).alignment = Alignment(wrap_text=True)

# ─────────────────────────────────────────────── Board Calendar
ws = sheet("Board Calendar", "Board Calendar — meetings, the statutory gap, and what the CEO personally signs",
  "Fills C17 and C18. Every section number, form and due date below is VERIFY LIVE — confirm with the Company Secretary before relying on it. The COO's Compliance Calendar owns the filing machinery; this sheet owns the board's own trail.",
  [8, 14, 16, 14, 14, 14, 14, 16, 40])
band(ws, 4, "A. Board meetings — Companies Act 2013 §173 (VERIFY LIVE with the CS)", 9)
header(ws, 5, ["#", "Date", "Type", "Gap from previous (days)", "Notice sent", "Pack circulated", "Quorum met", "Minutes signed on", "Key resolutions / decisions taken"])
ws.freeze_panes = "B6"
r = 6
meetings = [(date(2026,5,20),"Q1 board"), (date(2026,8,19),"Q2 board"), (date(2026,11,18),"Q3 board"), (date(2027,2,17),"Q4 board + accounts")]
for i, (d, t) in enumerate(meetings, start=1):
    gap = "" if i == 1 else f"=B{r}-B{r-1}"
    put(ws, r, [i, d, t, gap, "", "", "", "", ""], example=True,
        fmt={2: "dd mmm yyyy", 5: "dd mmm yyyy", 6: "dd mmm yyyy", 8: "dd mmm yyyy"},
        formula_cols=(4,) if i > 1 else (), input_cols=(5,6,7,8,9))
    r += 1
for rr in range(r, r + 4):
    put(ws, rr, ["", "", "", f"=IF(OR(B{rr}=\"\",B{rr-1}=\"\"),\"\",B{rr}-B{rr-1})", "", "", "", "", ""],
        fmt={2: "dd mmm yyyy", 5: "dd mmm yyyy", 6: "dd mmm yyyy", 8: "dd mmm yyyy"}, formula_cols=(4,), input_cols=(1,2,3,5,6,7,8,9))
ws.conditional_formatting.add(f"D6:D{r+3}", CellIsRule(operator="greaterThan", formula=["120"], fill=PatternFill("solid", fgColor="FEE2E2")))
r += 5
ws.cell(row=r-1, column=9, value="Red gap = more than 120 days between meetings. Confirm the exact statutory limit and any small-company relaxation with the CS — VERIFY LIVE.").font = Font(sz=9, i=True, color=MUTED)
band(ws, r, "B. The annual trail the CEO personally signs or tables", 9); r += 1
header(ws, r, ["#", "Item", "Form (VERIFY LIVE)", "Statutory due (VERIFY LIVE)", "Our target date", "Owner", "Status", "Evidence / file", "Note"]); r += 1
items = [("Disclosure of interest by every director", "MBP-1", "First board meeting of the FY, and on any change", "CS", "Critical in a family company — relatives' firms count as related parties."),
         ("Declaration of non-disqualification", "DIR-8", "First board meeting of the FY", "CS", ""),
         ("Register of contracts in which directors are interested", "MBP-4 / register", "Maintained continuously", "CS", "Any transaction with a relative's firm goes here before it happens, not after."),
         ("Board's report and financial statements approved", "Board resolution", "Before the AGM", "CEO + CA", "The CEO signs. Read them first — signing without reading is the failure."),
         ("Annual General Meeting", "—", "Within the statutory window after FY end", "CS", ""),
         ("Financial statements filed", "AOC-4", "After the AGM", "CS + CA", ""),
         ("Annual return filed", "MGT-7 / MGT-7A", "After the AGM", "CS", "Small companies may use the abridged form — confirm eligibility."),
         ("Director KYC", "DIR-3 KYC", "Annual, per director", "Each director personally", "Missing this deactivates the DIN. It is the cheapest way to embarrass yourself."),
         ("Return of deposits", "DPT-3", "Annual", "CS + CA", "Catches director loans — relevant if founders fund the company."),
         ("MSME payment return", "MSME-1", "Half-yearly", "CS + CA", "Applies to dues outstanding to MSME suppliers."),
         ("Auditor appointment / ratification", "ADT-1", "On appointment", "CS", ""),
         ("Statutory registers up to date", "Registers", "Continuous", "CS", "A future investor's diligence starts here.")]
for i, (item, form, due, own, note) in enumerate(items, start=1):
    put(ws, r, [i, item, form, due, "", own, "Not started", "", note], example=True, fmt={5: "dd mmm yyyy"}, input_cols=(5,7,8)); r += 1
r += 1
band(ws, r, "C. Advisory board — the outside voice a family board most needs", 9); r += 1
header(ws, r, ["#", "Name", "Why them", "Ask", "Offer", "Cadence", "Approached on", "Status", "Standing agenda slot"]); r += 1
adv = [("Fuel-industry veteran (ex-OMC or large dealer)", "Knows the buyer better than we do", "2 hours/month + intros", "0.1–0.25% over 2 years, quarterly vest — VERIFY structure with CS", "Monthly call"),
       ("Indian B2B SaaS operator", "Has sold to SMBs at this ticket size", "Quarterly review of the strategy kernel", "0.1–0.25% over 2 years", "Quarterly"),
       ("Independent finance person (not our CA)", "Reads the numbers with no stake in the story", "Attend two board meetings a year", "Honorarium", "Half-yearly")]
for i, (n, why, ask, offer, cad) in enumerate(adv, start=1):
    put(ws, r, [i, n, why, ask, offer, cad, "", "Not approached", "Yes"], example=True, fmt={7: "dd mmm yyyy"}, input_cols=(2,7,8)); r += 1

# ─────────────────────────────────────────────── Pipeline
ws = sheet("Pipeline", "Pipeline — the CEO's own founder-led sales pipeline",
  "Fills C10. This is the CEO's personal list, not the CRM — the COO course owns CRM hygiene and stage definitions. The only two columns that matter are Next step and Next date; a blank pair means the deal is dead and you have not admitted it.",
  [24, 14, 9, 20, 18, 16, 13, 34, 13, 14, 11, 14, 26, 22])
header(ws, 4, ["Dealer / firm", "District", "GSTINs", "Contact + role", "Source", "Stage", "Stage since", "Next step (a verb and a name)", "Next date", "₹ ARR potential", "Prob.", "Weighted ₹", "The objection actually blocking this", "Loss reason (if lost)"])
ws.freeze_panes = "B5"
stages = ["1 Identified","2 Conversation held","3 Demo done","4 Pilot agreed","5 Pilot running","6 Verbal yes","7 Paid","X Lost"]
r = 5
put(ws, r, ["Example Fuels, Raipur", "Raipur", 2, "Owner, 52", "Dealer association meet", "3 Demo done", date(2026,8,18),
            "Call owner's accountant with the Tally export sample — CEO", date(2026,8,28), 60000, 0.3, f"=IF(OR(J{r}=\"\",K{r}=\"\"),\"\",J{r}*K{r})",
            "\"My accountant already does this in Tally\"", ""],
    example=True, fmt={7: "dd mmm yyyy", 9: "dd mmm yyyy", 10: RS, 11: PCT, 12: RS}, formula_cols=(12,))
for rr in range(6, 46):
    put(ws, rr, ["", "", "", "", "", "", "", "", "", "", "", f"=IF(OR(J{rr}=\"\",K{rr}=\"\"),\"\",J{rr}*K{rr})", "", ""],
        fmt={7: "dd mmm yyyy", 9: "dd mmm yyyy", 10: RS, 11: PCT, 12: RS}, formula_cols=(12,),
        input_cols=(1,2,3,4,5,6,7,8,9,10,11,13,14))
ws.conditional_formatting.add("I5:I45", CellIsRule(operator="lessThan", formula=["TODAY()"], fill=PatternFill("solid", fgColor="FEE2E2")))
r = 47
band(ws, r, "Summary", 14); r += 1
put(ws, r, ["Open deals", "=COUNTIFS(F5:F45,\"<>\",F5:F45,\"<>X Lost\")", None, "Weighted pipeline ₹", "=SUMIFS(L5:L45,F5:F45,\"<>X Lost\")", None, "Overdue next steps", "=COUNTIFS(I5:I45,\"<\"&TODAY(),F5:F45,\"<>X Lost\")"],
    formula_cols=(2,5,8), fmt={5: RS})
r += 2
band(ws, r, "Objection bank — every objection heard, and the answer that worked", 14); r += 1
header(ws, r, ["Objection (in the dealer's words)", "Times heard", "What is really being said", "The answer that worked", "Proof to show", "", "", "", "", "", "", "", "", ""]); r += 1
obj = [("\"My accountant already does this in Tally\"", 8, "I don't want a second system, and my accountant is my trusted adviser", "We don't replace Tally, we feed it. Your accountant gets a cleaner export and stops chasing you for missing slips.", "A Tally-ready export from a real dealer's month"),
       ("\"My customers will never use an app\"", 7, "I'll end up doing double work", "The dealer-side works alone. Customer confirmations are a bonus, and we onboard your top 5 customers ourselves.", "Confirmation rate from an existing dealer's customers"),
       ("\"What if you shut down? My data is in there.\"", 6, "You're three people in Raipur", "One-click full export any day, in your name, and a written data-return clause in the contract.", "The export, run live on his phone"),
       ("\"It's too costly\"", 9, "I can't see the return yet", "Show his own reconciliation hours and last quarter's credit leakage. Price against that, not against ₹0.", "His own numbers, on paper"),
       ("\"Send me details on WhatsApp, I'll see\"", 12, "Polite no", "Ask for a 20-minute slot at his pump at his quietest hour, with his accountant present.", "A calendar invite, before you leave the room")]
for o, n, real, ans, proof in obj:
    put(ws, r, [o, n, real, ans, proof], example=True); r += 1

# ─────────────────────────────────────────────── Comp Bands
ws = sheet("Comp Bands", "Comp Bands — salary and equity by role and level, plus the offer check",
  "Fills C20. Bands are illustrative and MUST be re-benchmarked before any offer — VERIFY LIVE. Publishing a band internally is a culture decision, not a payroll one; decide it deliberately (lesson 11).",
  [26, 10, 16, 16, 16, 14, 14, 14, 36])
band(ws, 4, "A. Bands (annual CTC ₹, Raipur / remote-India)", 9)
header(ws, 5, ["Role family", "Level", "Min ₹", "Mid ₹", "Max ₹", "Equity %", "Options", "Location factor", "Note"])
ws.freeze_panes = "A6"
r = 6
bands = [("Engineering","L3",400000,550000,700000,0.001,"Junior / 0–2 yrs"),
         ("Engineering","L4",700000,950000,1200000,0.003,"Mid / 2–5 yrs"),
         ("Engineering","L5",1200000,1600000,2100000,0.006,"Senior / lead"),
         ("Engineering","L6",2000000,2600000,3400000,0.015,"CTO / function head — expect to pay above band or pay in equity"),
         ("Implementation / Support","L3",250000,340000,430000,0.001,"Field-capable; this is the role that decides whether onboarding scales"),
         ("Implementation / Support","L4",430000,560000,700000,0.002,""),
         ("Sales","L4",400000,550000,700000,0.002,"Plus variable — set the variable at 30–40% of target CTC"),
         ("Sales","L6",1200000,1600000,2200000,0.012,"First sales leader; do not hire before the script exists"),
         ("Founder-Director","—",600000,900000,1200000,0.0,"Directors' remuneration for a Pvt Ltd needs a board resolution — confirm with the CS. Pay yourself something.")]
for fam, lvl, mn, md, mx, eq, note in bands:
    put(ws, r, [fam, lvl, mn, md, mx, eq, f"=ROUND(F{r}*'ESOP Pool'!$B$7/'ESOP Pool'!$B$6,0)", 1.0, note],
        example=True, fmt={3: RS, 4: RS, 5: RS, 6: PCT, 7: "#,##0", 8: "0.00"}, formula_cols=(7,), input_cols=(3,4,5,6,8))
    r += 1
r += 1
band(ws, r, "B. Offer check — fill before the offer goes out", 9); r += 1
header(ws, r, ["Field", "Value", "", "", "", "", "", "", "Check"]); r += 1
o = r
offer = [("Candidate", "A. Example"), ("Role family", "Engineering"), ("Level", "L4"),
         ("Proposed CTC (₹)", 1000000), ("Proposed options", 300), ("Band mid for this level (₹)", "=SUMPRODUCT((A6:A14=B"+str(o+1)+")*(B6:B14=B"+str(o+2)+")*D6:D14)")]
for n, v in offer:
    isf = isinstance(v, str) and v.startswith("=")
    put(ws, r, [n, v], fmt={2: RS if "₹" in n else "General"}, formula_cols=(2,) if isf else (), input_cols=() if isf else (2,))
    if not isf: ws.cell(row=r, column=2).font = Font(i=True, color="808080"); ws.cell(row=r,column=2).fill = EX_FILL
    r += 1
put(ws, r, ["Inside band?", f'=IF(B{o+5}="","",IF(AND(B{o+3}>=B{o+5}*0.75,B{o+3}<=B{o+5}*1.3),"Yes","OUTSIDE BAND — write the reason below and get a second opinion"))',
            None,None,None,None,None,None, "An outside-band offer is allowed. An undocumented one is not."], formula_cols=(2,)); r += 1
put(ws, r, ["Reason if outside band", ""], input_cols=(2,)); r += 1
put(ws, r, ["Would we pay this to an existing person doing the same job?", ""], input_cols=(2,))
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
r += 2
ws.cell(row=r, column=1, value="Family-fairness rule (lesson 11): a related employee is paid from the same band, evaluated against criteria written before the review, and gets more scrutiny than anyone else — not less. Write the deviation down or do not make it.").font = Font(sz=9, i=True, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True); ws.row_dimensions[r].height = 30

# ─────────────────────────────────────────────── Risk & Pre-Mortem
ws = sheet("Risk & Pre-Mortem", "Risk & Pre-Mortem — it is 24 months from now and VSYST failed. Why?",
  "Fills C30. The COO workbook's Risk Register owns operational risk; this sheet owns the existential kind. Run the pre-mortem with all three founders in the room, silently and in writing first, then compare — that is what stops the loudest voice setting the list.",
  [8, 40, 12, 10, 10, 36, 34, 14, 14, 20])
band(ws, 4, "A. Pre-mortem — causes of death, scored", 10)
header(ws, 5, ["#", "Cause of failure", "Likelihood 1-5", "Impact 1-5", "Score", "The early-warning signal we would actually see", "Mitigation (a dated action, not an intention)", "Owner", "By when", "Status"])
ws.freeze_panes = "B6"
r = 6
risks = [("We run out of cash before the district motion is proven", 3, 5, "Runway falls below 9 months on the CEO Dashboard", "Trigger ladder in C13 pre-committed; annual prepay pricing; DPIIT/80-IAC and cloud credits applied for this quarter", "CEO", date(2026,10,31)),
 ("Dealers sign up and never reach a first invoice", 4, 5, "Median days-to-first-invoice rising above 14; activation below 50%", "Onboarding treated as the product; time-to-first-invoice on the dashboard; onboarding SOP owned by the COO", "COO", date(2026,11,30)),
 ("A ledger or billing bug misstates a dealer's money and destroys trust", 3, 5, "Any money-affecting incident; a reconciliation complaint", "Ledger correctness above features; the month-rebuild and check tooling; incident postmortem every time; CEO informed same day", "CEO", date(2026,9,30)),
 ("The founder-CEO stays the only salesperson and the motion never becomes repeatable", 4, 4, "No written script after 20 conversations; CAC not falling; every deal needs the CEO", "C10 script and objection bank; first sales hire trigger written down; CEO tracks % of deals not personally closed", "CEO", date(2026,12,31)),
 ("A family disagreement becomes a company disagreement", 3, 5, "A decision reopened twice; a founder silent in a meeting; a conversation that only happens at home", "C02 founders' agreement with vesting; C03 decision rights; the disagreement protocol; one outside advisor with a standing slot", "CEO", date(2026,10,15)),
 ("A larger incumbent or an OMC-backed vendor ships an adequate free version", 2, 5, "An OMC circular; a dealer mentioning a bundled tool; an ATG vendor adding billing", "Switching-cost moat: ledger, rate history and audit trail; deepen the two-sided customer link; move faster in-district", "CEO", date(2027,3,31)),
 ("A data breach exposes dealer financial data", 2, 5, "Any unauthorised access; a tenant seeing another tenant's data", "Tenant isolation audit; DPDP readiness; ISO 27001 roadmap; cyber liability cover priced", "CEO + COO", date(2027,3,31)),
 ("We chase the OMC prize for a year and it converts nothing", 3, 4, "Two quarters of meetings with no written next step or empanelment path", "A written kill-criterion and a one-founder-quarter budget; C26 scoring before any further pursuit", "CEO", date(2026,11,30)),
 ("The CEO burns out or a founder becomes unavailable", 3, 5, "Working every Sunday for a month; no holiday in a year; decisions queuing", "The COO course's autopilot test; a named backup for every founder-held relationship; a real holiday booked", "CEO", date(2027,1,31)),
 ("We stay pre-revenue because we keep building instead of selling", 4, 5, "Shipping features while the pipeline is empty; more commits than conversations this month", "Founder conversations on the CEO Dashboard; the no-list in C25; a weekly minimum of customer conversations", "CEO", date(2026,9,15))]
first_risk = r
for i, (c, l, im, sig, mit, own, by) in enumerate(risks, start=1):
    put(ws, r, [i, c, l, im, f"=C{r}*D{r}", sig, mit, own, by, "Open"], example=True,
        fmt={5: "0", 9: "dd mmm yyyy"}, formula_cols=(5,), input_cols=(3,4,10))
    ws.row_dimensions[r].height = 40
    r += 1
ws.conditional_formatting.add(f"E{first_risk}:E{r-1}", ColorScaleRule(
    start_type="num", start_value=1, start_color="DCFCE7", mid_type="num", mid_value=12, mid_color="FEF3C7",
    end_type="num", end_value=25, end_color="FEE2E2"))
r += 1
band(ws, r, "B. Crisis playbooks — written before the crisis, not during", 10); r += 1
header(ws, r, ["#", "Crisis", "First hour: who does what", "", "", "Who speaks, and what the first statement says", "Notification obligations (VERIFY LIVE with counsel)", "Playbook owner", "Last rehearsed", "Status"]); r += 1
cr = [("Dealer financial data breach", "CEO informed immediately; access revoked; scope determined before any statement", "CEO only. State what is known, what is not, what we are doing, when we will update next. Never speculate on cause.", "DPDP Act 2023 obligations to the Data Protection Board and affected persons — confirm timing with counsel", "CEO"),
 ("Ledger error that misstates money", "Freeze the affected postings; quantify the blast radius per tenant before contacting anyone", "CEO, to each affected dealer personally and by phone first. Exact rupee impact, the correction, and the date it is fixed.", "Contractual; possibly GST implications if invoices were affected — confirm with the CA", "CEO"),
 ("Extended outage during the 22:00–06:00 rate-confirmation window", "COO runs the incident; CEO handles dealers; fall back to a manual confirmation record", "COO for status, CEO for anything that touches a dealer's money or trust", "None statutory; contractual SLA if any", "COO"),
 ("Founder or partner dispute becomes public", "No response for 24 hours except an acknowledgement; counsel engaged", "CEO, one statement, agreed with counsel and co-founders. Never respond in a comment thread.", "Depends — take advice before saying anything", "CEO"),
 ("Key person unavailable (illness, accident)", "Named backup takes each relationship per the relationship map (lesson 14)", "CEO or the designated backup, to team first, then customers", "Board informed; check bank/DSC signatory continuity", "CEO")]
for i, (name, hour, speak, notif, own) in enumerate(cr, start=1):
    put(ws, r, [i, name, hour, None, None, speak, notif, own, "", "Not rehearsed"], example=True, fmt={9: "dd mmm yyyy"}, input_cols=(9,10))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 46
    r += 1

wb.save("/Users/shikhar/Documents/KIT/GITHUB/DZZLO_OMS/v1_79/obsidian-notes/content/vsyst-technologies/sakuradocs/CEO-Docs/toolkit/vsyst-ceo-workbook.xlsx")
print("OK — sheets:", wb.sheetnames)
